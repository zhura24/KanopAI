from pathlib import Path

import numpy as np
import rasterio
from rasterio.transform import from_origin
from openpyxl import load_workbook

from core.inference_engine import export_result_excel


def test_export_result_excel_writes_xlsx(tmp_path):
    raster_path = tmp_path / "sample.tif"
    with rasterio.open(
        raster_path,
        "w",
        driver="GTiff",
        height=4,
        width=4,
        count=2,
        dtype="uint16",
        transform=from_origin(0, 4, 1, 1),
    ) as dst:
        dst.write(np.array([[1, 2, 3, 4], [5, 6, 7, 8]], dtype=np.uint16), 1)
        dst.write(np.array([[9, 10, 11, 12], [13, 14, 15, 16]], dtype=np.uint16), 2)

    out_path = tmp_path / "result.xlsx"
    saved_path = export_result_excel(
        raster_path,
        out_path,
        np.array([[0.0, 0.0, 3.0, 3.0]], dtype=np.float32),
        np.array([0.95], dtype=np.float32),
        np.array([0], dtype=np.int32),
        class_names=["palm"],
    )

    assert saved_path == out_path
    assert out_path.exists()
    assert out_path.stat().st_size > 0


def test_export_result_excel_uses_polygon_mean_and_excludes_pixels(tmp_path):
    raster_path = tmp_path / "sample.tif"
    with rasterio.open(
        raster_path,
        "w",
        driver="GTiff",
        height=4,
        width=4,
        count=1,
        dtype="uint16",
        transform=from_origin(0, 4, 1, 1),
    ) as dst:
        dst.write(np.arange(1, 17, dtype=np.uint16).reshape(4, 4), 1)

    out_path = tmp_path / "polygon-result.xlsx"
    export_result_excel(
        raster_path,
        out_path,
        np.array([[0.0, 0.0, 4.0, 4.0]], dtype=np.float32),
        np.array([0.95], dtype=np.float32),
        np.array([0], dtype=np.int32),
        aoi_polygons_px=[[(0, 0), (4, 0), (4, 4), (0, 4)]],
        exclude_polygons_px=[[(0, 0), (2, 0), (2, 4), (0, 4)]],
    )

    sheet = load_workbook(out_path, data_only=True).active
    assert sheet.cell(2, 4).value == 9.5
