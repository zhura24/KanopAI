from pathlib import Path

import numpy as np
import rasterio
import shapefile

from core.inference_engine import (
    load_inference_result_from_shapefile,
    compute_visible_crop_region,
    export_result_excel,
    auto_detect_band_mapping_multiref,
)


def test_compute_visible_crop_region_clamps_to_raster_bounds():
    bounds = compute_visible_crop_region(2000, 1500, (-50, -20, 250, 1200))
    assert bounds == (0, 0, 250, 1200)


def test_load_inference_result_from_shapefile(tmp_path):
    shp_path = tmp_path / "inference_result.shp"

    with shapefile.Writer(str(shp_path), shapeType=shapefile.POLYGON) as writer:
        writer.field("id", "N", size=10)
        writer.field("kelas", "C", size=20)
        writer.field("confidence", "N", size=10, decimal=4)
        writer.field("model", "C", size=30)
        writer.field("x1_px", "N", size=10, decimal=1)
        writer.field("y1_px", "N", size=10, decimal=1)
        writer.field("x2_px", "N", size=10, decimal=1)
        writer.field("y2_px", "N", size=10, decimal=1)
        writer.field("status", "C", size=20)

        polygon = [[0.0, 0.0], [100.0, 0.0], [100.0, 50.0], [0.0, 50.0], [0.0, 0.0]]
        writer.poly([polygon])
        writer.record(1, "palm", 0.92, "demo", 0.0, 0.0, 100.0, 50.0, "not_corrected")

    result = load_inference_result_from_shapefile(shp_path)

    assert result is not None
    assert result.boxes.shape == (1, 4)
    assert result.scores.shape == (1,)
    assert result.classes.shape == (1,)
    assert result.class_names == ["palm"]


def test_load_inference_result_from_georeferenced_shape_uses_geometry_when_fields_are_stale(tmp_path):
    shp_path = tmp_path / "inference_result.shp"
    raster_path = tmp_path / "demo.tif"

    transform = rasterio.transform.from_origin(100.0, 200.0, 1.0, 1.0)
    height = 400
    width = 500
    data = np.zeros((height, width), dtype=np.uint8)
    with rasterio.open(
        raster_path,
        "w",
        driver="GTiff",
        height=height,
        width=width,
        count=1,
        dtype="uint8",
        crs="EPSG:4326",
        transform=transform,
    ) as dst:
        dst.write(data, 1)

    with shapefile.Writer(str(shp_path), shapeType=shapefile.POLYGON) as writer:
        writer.field("id", "N", size=10)
        writer.field("kelas", "C", size=20)
        writer.field("confidence", "N", size=10, decimal=4)
        writer.field("model", "C", size=30)
        writer.field("x1_px", "N", size=10, decimal=1)
        writer.field("y1_px", "N", size=10, decimal=1)
        writer.field("x2_px", "N", size=10, decimal=1)
        writer.field("y2_px", "N", size=10, decimal=1)
        writer.field("status", "C", size=20)

        polygon = [
            [100.0, 198.0],
            [102.0, 198.0],
            [102.0, 195.0],
            [100.0, 195.0],
            [100.0, 198.0],
        ]
        writer.poly([polygon])
        writer.record(1, "palm", 0.77, "demo", 9999.0, 9999.0, 10000.0, 10000.0, "not_corrected")

    result = load_inference_result_from_shapefile(shp_path, raster_path=raster_path)

    assert result.boxes.shape == (1, 4)
    assert result.boxes[0, 0] == 0.0
    assert result.boxes[0, 1] == 2.0
    assert result.boxes[0, 2] == 2.0
    assert result.boxes[0, 3] == 5.0
    assert result.class_names == ["palm"]


def test_auto_detect_band_mapping_multiref_prefers_best_remaining_input(tmp_path):
    raster_path = tmp_path / "demo_multiref.tif"
    transform = rasterio.transform.from_origin(100.0, 200.0, 1.0, 1.0)
    with rasterio.open(
        raster_path,
        "w",
        driver="GTiff",
        height=50,
        width=50,
        count=3,
        dtype="float32",
        crs="EPSG:4326",
        transform=transform,
    ) as dst:
        band1 = np.full((50, 50), 10, dtype=np.float32)
        band2 = np.full((50, 50), 20, dtype=np.float32)
        band3 = np.full((50, 50), 30, dtype=np.float32)
        dst.write(band1, 1)
        dst.write(band2, 2)
        dst.write(band3, 3)

    band_stats = {
        1: {"sources": {"ref_a": {"mean": 10.0, "p_low": 0, "p_high": 20}, "ref_b": {"mean": 60.0, "p_low": 0, "p_high": 20}}},
        2: {"sources": {"ref_a": {"mean": 20.0, "p_low": 0, "p_high": 30}, "ref_b": {"mean": 5.0, "p_low": 0, "p_high": 30}}},
        3: {"sources": {"ref_a": {"mean": 30.0, "p_low": 0, "p_high": 40}, "ref_b": {"mean": 40.0, "p_low": 0, "p_high": 40}}},
    }

    with rasterio.open(raster_path) as src:
        mapping = auto_detect_band_mapping_multiref(src, band_stats)

    assert mapping[1]["input_band"] == 1
    assert mapping[2]["input_band"] == 2
    assert mapping[3]["input_band"] == 3


def test_export_result_excel_fast_mode_writes_centroid_metrics(tmp_path):
    raster_path = tmp_path / "demo.tif"
    transform = rasterio.transform.from_origin(100.0, 200.0, 1.0, 1.0)
    with rasterio.open(
        raster_path,
        "w",
        driver="GTiff",
        height=10,
        width=10,
        count=1,
        dtype="uint8",
        crs="EPSG:4326",
        transform=transform,
    ) as dst:
        dst.write(np.zeros((10, 10), dtype=np.uint8), 1)

    out_path = tmp_path / "fast_export.xlsx"
    result = export_result_excel(
        str(raster_path),
        str(out_path),
        np.array([[0.0, 0.0, 2.0, 2.0]], dtype=np.float32),
        np.array([0.92], dtype=np.float32),
        np.array([0], dtype=np.int32),
        class_names=["palm"],
        fast_mode=True,
    )

    assert result is not None
    assert out_path.exists()

    from openpyxl import load_workbook
    ws = load_workbook(out_path).active
    assert ws[1][0].value == "Latitude"
    assert ws[1][1].value == "Longitude"
    assert ws[1][2].value == "radius_m"
    assert ws[1][3].value == "diameter_m"
    assert ws[1][4].value == "area_m2"
    assert ws[2][2].value > 0
    assert ws[2][3].value > 0
    assert ws[2][4].value > 0
