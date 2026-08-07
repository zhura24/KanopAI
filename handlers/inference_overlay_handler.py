"""Handler & Interactive Manager untuk Vector Overlay Bounding Box Inference.

Menyediakan rendering interaktif hasil deteksi YOLO multispektral pada RasterViewer,
toolset manual correction (Add Box, Edit/Resize Box, Delete Box), sistem Undo/Redo,
    serta ekspor dataset shapefile raw_detection.shp dan corrected_detection.shp.
"""

from typing import List, Dict, Any, Optional, Tuple
import logging
from pathlib import Path
from copy import deepcopy

from PyQt6.QtWidgets import QGraphicsRectItem, QGraphicsItem, QMessageBox, QFileDialog
from PyQt6.QtCore import Qt, QRectF, QPointF, pyqtSignal, QObject, QEvent
from PyQt6.QtGui import QPen, QBrush, QColor, QCursor

from core.inference_engine import (
    save_shapefile,
    save_corrected_shapefile,
    InferenceResult,
    resolve_class_name,
    _read_band_values_for_detection,
)


class ResizeHandle(QGraphicsRectItem):
    """Handle/Grip kecil di sudut bounding box untuk interaksi resize."""

    def __init__(self, position_name: str, parent: 'InferenceBoxItem') -> None:
        super().__init__(-4, -4, 8, 8, parent)
        self.position_name = position_name
        self.parent_box = parent

        self.setBrush(QBrush(QColor(255, 255, 255)))
        self.setPen(QPen(QColor(0, 0, 0), 1))
        self.setZValue(2005)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
        self.setAcceptHoverEvents(True)
        self.setCursor(self._get_cursor_for_position())
        self.setVisible(False)

    def _get_cursor_for_position(self) -> QCursor:
        if self.position_name in ('top_left', 'bottom_right'):
            return QCursor(Qt.CursorShape.SizeFDiagCursor)
        return QCursor(Qt.CursorShape.SizeBDiagCursor)


class InferenceBoxItem(QGraphicsRectItem):
    """Item QGraphicsRectItem interaktif untuk bounding box deteksi."""

    def __init__(self, box_id: int, box: List[float], score: float, class_name: str,
                 status: str = "retained", parent_handler: Optional['InferenceOverlayHandler'] = None) -> None:
        x1, y1, x2, y2 = box
        w = max(1.0, x2 - x1)
        h = max(1.0, y2 - y1)
        super().__init__(x1, y1, w, h)

        self.box_id = box_id
        self.score = score
        self.class_name = class_name
        self.status = status
        self.parent_handler = parent_handler

        self.setZValue(2000)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemSendsGeometryChanges, True)
        self.setAcceptedMouseButtons(Qt.MouseButton.NoButton)
        self.setAcceptHoverEvents(True)

        # PERF: handle resize (4x QGraphicsRectItem per box) SENGAJA tidak dibuat
        # di sini lagi. Dulu selalu dibuat di depan meski setVisible(False) --
        # untuk 5000 deteksi = 25.000 Qt item yang harus di-index scene sejak awal.
        # Sekarang dibuat lazy, baru saat box pertama kali di-select (lihat
        # apply_style()/_ensure_handles()). Behavior TIDAK berubah: handle memang
        # cuma pernah kelihatan & bisa diklik saat box ter-select di mode edit,
        # jadi menunda pembuatannya tidak menghilangkan interaksi apa pun.
        self.handles = {}
        self.apply_style()

    def _ensure_handles(self) -> None:
        """Buat 4 ResizeHandle sekali saja, on-demand, saat box pertama kali perlu tampilkan handle."""
        if self.handles:
            return
        self.handles = {
            'top_left': ResizeHandle('top_left', self),
            'top_right': ResizeHandle('top_right', self),
            'bottom_left': ResizeHandle('bottom_left', self),
            'bottom_right': ResizeHandle('bottom_right', self),
        }
        self.update_handle_positions()

    def update_handle_positions(self) -> None:
        if not self.handles:
            return
        r = self.rect()
        self.handles['top_left'].setPos(r.left(), r.top())
        self.handles['top_right'].setPos(r.right(), r.top())
        self.handles['bottom_left'].setPos(r.left(), r.bottom())
        self.handles['bottom_right'].setPos(r.right(), r.bottom())

    def get_box_coords(self) -> List[float]:
        r = self.rect()
        return [float(r.x()), float(r.y()), float(r.x() + r.width()), float(r.y() + r.height())]

    def apply_style(self) -> None:
        if self.status == "eliminated":
            pen_color = QColor(239, 68, 68, 180)
            fill_color = QColor(239, 68, 68, 30)
            pen_style = Qt.PenStyle.DashLine
        elif self.status == "added":
            pen_color = QColor(6, 182, 212, 230)
            fill_color = QColor(6, 182, 212, 50)
            pen_style = Qt.PenStyle.SolidLine
        elif self.status == "edited":
            pen_color = QColor(234, 179, 8, 230)
            fill_color = QColor(234, 179, 8, 50)
            pen_style = Qt.PenStyle.SolidLine
        else:
            pen_color = QColor(34, 197, 94, 230)
            fill_color = QColor(34, 197, 94, 40)
            pen_style = Qt.PenStyle.SolidLine

        width = 3 if self.isSelected() else 2
        self.setPen(QPen(pen_color, width, pen_style))
        self.setBrush(QBrush(fill_color))

        is_edit_active = self.parent_handler and self.parent_handler.current_mode == "edit"
        show_handles = self.isSelected() and is_edit_active and self.status != "eliminated"
        if show_handles and not self.handles:
            self._ensure_handles()
        for h in self.handles.values():
            h.setVisible(show_handles)

    def mousePressEvent(self, event):
        if self.parent_handler and self.parent_handler.current_mode == "edit" and event.button() == Qt.MouseButton.LeftButton:
            for bi in self.parent_handler.box_items:
                bi.setSelected(bi is self)
            self.setSelected(True)
            if self.parent_handler:
                self.parent_handler._on_item_selection_changed()
            super().mousePressEvent(event)
        else:
            event.ignore()

    def mouseReleaseEvent(self, event):
        if self.parent_handler and self.parent_handler.current_mode == "edit" and event.button() == Qt.MouseButton.LeftButton:
            super().mouseReleaseEvent(event)
        else:
            event.ignore()

    def itemChange(self, change: QGraphicsItem.GraphicsItemChange, value: Any) -> Any:
        if change == QGraphicsItem.GraphicsItemChange.ItemSelectedHasChanged:
            self.apply_style()
            if self.parent_handler:
                self.parent_handler._on_item_selection_changed()
        elif change == QGraphicsItem.GraphicsItemChange.ItemPositionHasChanged:
            self.update_handle_positions()
        return super().itemChange(change, value)


class InferenceUndoStack:
    """Pengelola Undo / Redo stack untuk pengeditan bounding box."""

    def __init__(self) -> None:
        self.history: List[Dict[str, Any]] = []
        self.index: int = -1

    def push_action(self, action: Dict[str, Any]) -> None:
        self.history = self.history[:self.index + 1]
        self.history.append(action)
        self.index += 1

    def can_undo(self) -> bool:
        return self.index >= 0

    def can_redo(self) -> bool:
        return self.index < len(self.history) - 1

    def advance(self) -> Optional[Dict[str, Any]]:
        if self.can_redo():
            self.index += 1
            return self.history[self.index]
        return None

    def retreat(self) -> Optional[Dict[str, Any]]:
        if self.can_undo():
            act = self.history[self.index]
            self.index -= 1
            return act
        return None

    def clear(self) -> None:
        self.history.clear()
        self.index = -1


class InferenceOverlayHandler(QObject):
    """Handler utama untuk overlay bounding box dan interaksi UI."""

    def __init__(self, main_window: Any) -> None:
        super().__init__()
        self.main_window = main_window
        self.logger = logging.getLogger(__name__)

        self.box_items: List[InferenceBoxItem] = []
        self.raw_boxes_backup: List[Dict[str, Any]] = []
        self.confidence_threshold: float = 0.0
        self._aoi_polygons_px = None
        self._exclude_polygons_px = None
        self.undo_stack = InferenceUndoStack()

        self.current_mode: str = "none"
        self._is_drawing: bool = False
        self._draw_start_pos: Optional[QPointF] = None
        self._temp_rubberband: Optional[QGraphicsRectItem] = None
        self._next_box_id: int = 1

        self._active_handle: Optional[ResizeHandle] = None
        self._drag_start_box_rect: Optional[QRectF] = None
        self._move_start_pos: Optional[QPointF] = None
        self._edit_old_box: Optional[List[float]] = None
        self._edit_old_status: Optional[str] = None
        self._last_result_signature: Optional[Tuple[Any, ...]] = None

        self._install_viewer_hooks()

    def _install_viewer_hooks(self) -> None:
        viewer = getattr(self.main_window, "viewer", None)
        if viewer and viewer.viewport():
            viewer.viewport().installEventFilter(self)

    def eventFilter(self, obj, event) -> bool:
        viewer = getattr(self.main_window, "viewer", None)
        if not viewer or obj is not viewer.viewport():
            return False

        if self.current_mode == "none":
            return False

        et = event.type()
        if et == QEvent.Type.MouseButtonPress and event.button() == Qt.MouseButton.LeftButton:
            return self._handle_mouse_press(viewer, event)
        if et == QEvent.Type.MouseMove:
            return self._handle_mouse_move(viewer, event)
        if et == QEvent.Type.MouseButtonRelease and event.button() == Qt.MouseButton.LeftButton:
            return self._handle_mouse_release(viewer, event)
        return False

    def _scene_pos(self, viewer, event) -> QPointF:
        return viewer.mapToScene(event.position().toPoint())

    def _handle_mouse_press(self, viewer, event) -> bool:
        scene_pos = self._scene_pos(viewer, event)

        if self.current_mode == "add":
            self._is_drawing = True
            self._draw_start_pos = scene_pos
            if self._temp_rubberband:
                viewer.scene.removeItem(self._temp_rubberband)
            self._temp_rubberband = QGraphicsRectItem(
                QRectF(scene_pos, scene_pos)
            )
            self._temp_rubberband.setPen(QPen(QColor(37, 99, 235), 2, Qt.PenStyle.DashLine))
            self._temp_rubberband.setBrush(QBrush(QColor(37, 99, 235, 40)))
            self._temp_rubberband.setZValue(2001)
            viewer.scene.addItem(self._temp_rubberband)
            event.accept()
            return True

        if self.current_mode == "edit":
            item = viewer.scene.itemAt(scene_pos, viewer.transform())
            while item and not isinstance(item, InferenceBoxItem):
                if isinstance(item, ResizeHandle):
                    self._active_handle = item
                    self._drag_start_box_rect = item.parent_box.rect()
                    self._edit_old_box = item.parent_box.get_box_coords()
                    self._edit_old_status = item.parent_box.status
                    event.accept()
                    return True
                item = item.parentItem() if hasattr(item, 'parentItem') else None

            if isinstance(item, InferenceBoxItem) and item.status != "eliminated":
                for bi in self.box_items:
                    bi.setSelected(bi is item)
                self._move_start_pos = scene_pos
                self._drag_start_box_rect = item.rect()
                self._edit_old_box = item.get_box_coords()
                self._edit_old_status = item.status
                event.accept()
                return True

            for bi in self.box_items:
                bi.setSelected(False)
            self._update_box_info(None)
            event.accept()
            return True

        return False

    def _handle_mouse_move(self, viewer, event) -> bool:
        scene_pos = self._scene_pos(viewer, event)

        if self.current_mode == "add" and self._is_drawing and self._temp_rubberband and self._draw_start_pos:
            rect = QRectF(self._draw_start_pos, scene_pos).normalized()
            self._temp_rubberband.setRect(rect)
            event.accept()
            return True

        if self.current_mode == "edit" and self._active_handle and self._drag_start_box_rect:
            parent = self._active_handle.parent_box
            r = parent.rect()
            pos = self._active_handle.position_name
            if pos == 'top_left':
                r.setTopLeft(scene_pos)
            elif pos == 'top_right':
                r.setTopRight(scene_pos)
            elif pos == 'bottom_left':
                r.setBottomLeft(scene_pos)
            elif pos == 'bottom_right':
                r.setBottomRight(scene_pos)
            r = r.normalized()
            if r.width() >= 2 and r.height() >= 2:
                parent.setRect(r)
                parent.update_handle_positions()
            event.accept()
            return True

        if self.current_mode == "edit" and self._move_start_pos is not None:
            selected = [i for i in self.box_items if i.isSelected() and i.status != "eliminated"]
            if selected:
                delta = scene_pos - self._move_start_pos
                item = selected[0]
                new_rect = self._drag_start_box_rect.translated(delta)
                item.setRect(new_rect)
                item.update_handle_positions()
            event.accept()
            return True

        return False

    def _handle_mouse_release(self, viewer, event) -> bool:
        scene_pos = self._scene_pos(viewer, event)

        if self.current_mode == "add" and self._is_drawing and self._draw_start_pos:
            rect = QRectF(self._draw_start_pos, scene_pos).normalized()
            if self._temp_rubberband:
                viewer.scene.removeItem(self._temp_rubberband)
                self._temp_rubberband = None
            self._is_drawing = False
            self._draw_start_pos = None

            if rect.width() >= 5 and rect.height() >= 5:
                box = [rect.x(), rect.y(), rect.x() + rect.width(), rect.y() + rect.height()]
                self._add_new_box(box)
            event.accept()
            return True

        if self.current_mode == "edit":
            if self._active_handle or self._move_start_pos is not None:
                selected = [i for i in self.box_items if i.isSelected() and i.status != "eliminated"]
                for item in selected:
                    new_box = item.get_box_coords()
                    if self._edit_old_box and new_box != self._edit_old_box:
                        item.status = "edited"
                        item.apply_style()
                        self.undo_stack.push_action({
                            "type": "edit",
                            "item": item,
                            "old_box": self._edit_old_box,
                            "new_box": new_box,
                            "old_status": self._edit_old_status or "retained",
                        })
                        self._update_panel_ui()
            self._active_handle = None
            self._move_start_pos = None
            self._drag_start_box_rect = None
            self._edit_old_box = None
            self._edit_old_status = None
            event.accept()
            return True

        return False

    def _add_new_box(self, box: List[float]) -> None:
        box_id = self._next_box_id
        self._next_box_id += 1
        item = InferenceBoxItem(
            box_id=box_id,
            box=box,
            score=1.0,
            class_name="manual",
            status="added",
            parent_handler=self,
        )
        self._apply_interaction_flags(item)
        self.main_window.viewer.scene.addItem(item)
        self.box_items.append(item)
        self.undo_stack.push_action({"type": "add", "item": item})
        self._update_panel_ui()
        self.logger.info(f"Added manual box #{box_id}")

    def _apply_interaction_flags(self, item: InferenceBoxItem) -> None:
        editable = self.current_mode == "edit"
        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable, editable)
        item.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable, False)
        item.setAcceptedMouseButtons(
            Qt.MouseButton.LeftButton if editable else Qt.MouseButton.NoButton
        )

    def clear_overlay(self) -> None:
        if hasattr(self.main_window, "viewer") and self.main_window.viewer:
            scene = self.main_window.viewer.scene
            for item in self.box_items:
                try:
                    scene.removeItem(item)
                except Exception:
                    pass
            if self._temp_rubberband:
                try:
                    scene.removeItem(self._temp_rubberband)
                except Exception:
                    pass
        self.box_items.clear()
        self.raw_boxes_backup.clear()
        self.undo_stack.clear()
        self._temp_rubberband = None
        self._next_box_id = 1
        self._update_panel_ui()
        self._update_box_info(None)

    def set_overlay_visibility(self, visible: bool) -> None:
        """Set visibility of all inference box items on the scene."""
        for item in self.box_items:
            try:
                if item and item.scene():
                    item.setVisible(visible)
            except Exception as e:
                self.logger.debug(f"Error setting box item visibility: {e}")

    def set_confidence_threshold(self, threshold: float) -> None:
        """Live-filter box items by confidence score, without re-running inference.

        Boxes with score < threshold are hidden (setVisible(False)); boxes with
        score >= threshold are shown again. Manually eliminated boxes
        (status == 'eliminated') always stay hidden regardless of score —
        a manual delete should never be un-hidden by moving this slider.
        Manually added boxes (status == 'added', score == 1.0) are effectively
        always shown since their score is the max possible.

        Because both display AND export (export_shapefiles -> active_box_records)
        key off `item.isVisible()`, this single call keeps the canvas and the
        shp/xlsx/geojson export in sync automatically.
        """
        self.confidence_threshold = float(threshold)
        for item in self.box_items:
            try:
                if item.status == "eliminated":
                    continue
                item.setVisible(item.score >= self.confidence_threshold)
            except Exception as e:
                self.logger.debug(f"Error applying confidence threshold to box #{getattr(item, 'box_id', '?')}: {e}")
        self._update_panel_ui()

    def display_results(self, result: InferenceResult, replace_existing: bool = True) -> None:
        if result is None or result.boxes is None or len(result.boxes) == 0:
            self.logger.info("No inference boxes to display.")
            return

        result_signature = (
            tuple(map(float, result.boxes.reshape(-1))) if getattr(result, 'boxes', None) is not None and len(result.boxes) else tuple()
        )
        if replace_existing:
            self.clear_overlay()
        elif self.box_items and self._last_result_signature == result_signature:
            self.logger.info("Skipping redraw for unchanged inference result")
            return
        elif self.box_items:
            self.logger.info("Preserving existing inference overlays during layer sync")

        scene = self.main_window.viewer.scene
        class_names = result.class_names
        self._aoi_polygons_px = getattr(result, "aoi_polygons_px", None)
        self._exclude_polygons_px = getattr(result, "exclude_polygons_px", None)

        if replace_existing:
            self.box_items = []
            self.raw_boxes_backup = []
            self.undo_stack.clear()

        for idx, (box, score, cls) in enumerate(zip(result.boxes, result.scores, result.classes), start=1):
            x1, y1, x2, y2 = [float(v) for v in box]
            cls_name = resolve_class_name(int(cls), class_names)
            item = InferenceBoxItem(
                box_id=idx,
                box=[x1, y1, x2, y2],
                score=float(score),
                class_name=cls_name,
                status="retained",
                parent_handler=self,
            )
            self._apply_interaction_flags(item)
            scene.addItem(item)
            self.box_items.append(item)

            self.raw_boxes_backup.append({
                "id": idx,
                "box": [x1, y1, x2, y2],
                "score": float(score),
                "class": cls_name,
                "class_id": int(cls),
            })

        self._next_box_id = len(self.box_items) + 1
        self._last_result_signature = result_signature

        # Apply whatever confidence threshold is currently set (e.g. carried
        # over from the panel's spin_confidence_filter) to the freshly created
        # boxes, so canvas + export stay consistent with the UI slider state.
        if self.confidence_threshold > 0.0:
            self.set_confidence_threshold(self.confidence_threshold)

        # Synchronize inference result to active layer dictionary for persistence
        if hasattr(self.main_window, '_get_active_layer'):
            active_layer = self.main_window._get_active_layer()
            if active_layer is not None:
                active_layer['inference_result'] = result
                active_layer['detections'] = result

        self.logger.info(f"Rendered {len(self.box_items)} interactive bounding boxes on RasterViewer.")
        self._update_panel_ui()

    def set_mode(self, mode: str) -> None:
        self.current_mode = mode
        viewer = getattr(self.main_window, "viewer", None)
        if not viewer:
            return

        if mode == "add":
            viewer.setDragMode(viewer.DragMode.NoDrag)
            viewer.setCursor(Qt.CursorShape.CrossCursor)
        elif mode == "edit":
            viewer.setDragMode(viewer.DragMode.NoDrag)
            viewer.setCursor(Qt.CursorShape.ArrowCursor)
        else:
            viewer.setDragMode(viewer.DragMode.ScrollHandDrag)
            viewer.setCursor(Qt.CursorShape.OpenHandCursor)
            for item in self.box_items:
                item.setSelected(False)

        for item in self.box_items:
            self._apply_interaction_flags(item)
            item.apply_style()

        if mode == "none":
            self._update_box_info(None)

    def _on_item_selection_changed(self) -> None:
        panel = getattr(self.main_window, "inference_panel", None)
        if panel:
            has_sel = any(item.isSelected() for item in self.box_items)
            panel.btn_delete_box.setEnabled(has_sel)

        selected = [i for i in self.box_items if i.isSelected() and i.status != "eliminated"]
        self._update_box_info(selected[0] if selected else None)

    def _update_box_info(self, item: Optional[InferenceBoxItem]) -> None:
        panel = getattr(self.main_window, "inference_panel", None)
        if not panel or not hasattr(panel, "lbl_box_info"):
            return

        if item is None:
            panel.lbl_box_info.setText("Box Info: Select 'Edit' then click a box")
            return

        coords = item.get_box_coords()
        w = coords[2] - coords[0]
        h = coords[3] - coords[1]
        panel.lbl_box_info.setText(
            f"Box #{item.box_id} | Class: {item.class_name} | "
            f"Confidence: {item.score:.3f} | Status: {item.status} | "
            f"Size: {w:.0f}×{h:.0f}px | "
            f"Pos: ({coords[0]:.0f}, {coords[1]:.0f}) → ({coords[2]:.0f}, {coords[3]:.0f})"
        )

    def delete_selected_box(self) -> None:
        selected = [item for item in self.box_items if item.isSelected()]
        if not selected:
            return

        for item in selected:
            if item.status == "eliminated":
                continue
            old_status = item.status
            item.status = "eliminated"
            item.setVisible(False)
            item.apply_style()
            self.undo_stack.push_action({
                "type": "delete",
                "item": item,
                "old_status": old_status,
            })

        self._update_panel_ui()
        self._update_box_info(None)

    def undo_action(self) -> None:
        act = self.undo_stack.retreat()
        if not act:
            return

        t = act["type"]
        item = act["item"]

        if t == "delete":
            item.status = act["old_status"]
            item.setVisible(True)
            item.apply_style()
        elif t == "add":
            item.setVisible(False)
        elif t == "edit":
            old_box = act["old_box"]
            item.setRect(old_box[0], old_box[1], old_box[2] - old_box[0], old_box[3] - old_box[1])
            item.status = act["old_status"]
            item.update_handle_positions()
            item.apply_style()

        self._update_panel_ui()

    def redo_action(self) -> None:
        act = self.undo_stack.advance()
        if not act:
            return

        t = act["type"]
        item = act["item"]

        if t == "delete":
            item.status = "eliminated"
            item.setVisible(False)
            item.apply_style()
        elif t == "add":
            item.setVisible(True)
            item.status = "added"
            item.apply_style()
        elif t == "edit":
            new_box = act["new_box"]
            item.setRect(new_box[0], new_box[1], new_box[2] - new_box[0], new_box[3] - new_box[1])
            item.status = "edited"
            item.update_handle_positions()
            item.apply_style()

        self._update_panel_ui()

    def _update_panel_ui(self) -> None:
        panel = getattr(self.main_window, "inference_panel", None)
        if not panel:
            return

        raw_count = len(self.raw_boxes_backup)
        active_items = [it for it in self.box_items if it.status != "eliminated" and it.isVisible()]
        panel.lbl_summary.setText(
            f"Raw: {raw_count} | Active: {len(active_items)} boxes "
            f"(conf >= {self.confidence_threshold:.2f})"
        )
        panel.update_undo_redo_states(self.undo_stack.can_undo(), self.undo_stack.can_redo())

    def export_shapefiles(
        self,
        reviewer_name: str = "",
        correction_date: str = "",
    ) -> None:
        """Export inference detections from the active raster.

        The export supports a custom output directory and optional output stem.
        Saved files are named using the chosen stem or the raster file stem:
        - <stem>_raw_detection.shp
        - <stem>_corrected_detection.shp
        - <stem>_centroid.xlsx
        - <stem>_centroid.geojson
        """
        mw = self.main_window
        panel = getattr(mw, "inference_panel", None)
        raster_path = panel._get_active_raster_path() if panel else None
        if not raster_path and hasattr(mw, "_get_active_raster_path"):
            raster_path = mw._get_active_raster_path()

        if not raster_path or not Path(raster_path).exists():
            QMessageBox.critical(mw, "Export Error", "No active raster file available for export.")
            return

        p_raster = Path(raster_path)
        output_dir_text = None
        output_name_text = None
        if panel:
            output_dir_text = panel.txt_output_dir.text().strip() or None
            output_name_text = panel.txt_output_name.text().strip() or None

        out_dir = Path(output_dir_text) if output_dir_text else p_raster.parent / f"export_{p_raster.stem}"
        out_dir.mkdir(parents=True, exist_ok=True)

        output_stem = Path(output_name_text).stem if output_name_text else p_raster.stem
        raw_shp = out_dir / f"{output_stem}_raw_detection.shp"
        corrected_shp = out_dir / f"{output_stem}_corrected_detection.shp"

        import numpy as np
        if self.raw_boxes_backup:
            raw_boxes = np.array([b["box"] for b in self.raw_boxes_backup], dtype=np.float32)
            raw_scores = np.array([b["score"] for b in self.raw_boxes_backup], dtype=np.float32)
            raw_classes = np.array([b.get("class_id", 0) for b in self.raw_boxes_backup], dtype=np.int32)
            class_names = {b.get("class_id", 0): b["class"] for b in self.raw_boxes_backup}

            save_shapefile(p_raster, raw_boxes, raw_scores, raw_classes, raw_shp,
                           model_name="yolo_multispectral", class_names=class_names)

        corr_boxes, corr_scores, corr_classes, corr_statuses = [], [], [], []
        class_names_corr = {}
        active_box_records = []

        # NOTE: corrected_detection.shp / .xlsx / .geojson all follow the
        # live Confidence Filter (spin_confidence_filter). A box is included
        # here only if it is currently visible on the canvas, i.e. it was not
        # manually deleted (status == 'eliminated') AND its score is >= the
        # current confidence threshold. raw_detection.shp above is the only
        # export that stays a full, unfiltered audit dump of everything the
        # model produced.
        for item in self.box_items:
            if item.status == "eliminated" or not item.isVisible():
                continue
            b_coords = item.get_box_coords()
            corr_boxes.append(b_coords)
            corr_scores.append(item.score)
            corr_classes.append(0)
            corr_statuses.append(item.status)
            class_names_corr[0] = item.class_name
            active_box_records.append({
                "id": item.box_id,
                "box": b_coords,
                "class_name": item.class_name,
            })

        # Fallback only applies when box_items itself was never populated
        # (e.g. imported session edge case). It must NOT trigger just because
        # the confidence filter legitimately excluded everything -- otherwise
        # raising the threshold to "nothing passes" would silently dump the
        # full raw set again, defeating the filter.
        if not active_box_records and not self.box_items and self.raw_boxes_backup:
            for idx, raw_box in enumerate(self.raw_boxes_backup, start=1):
                active_box_records.append({
                    "id": raw_box.get("id", idx),
                    "box": raw_box["box"],
                    "class_name": raw_box.get("class", "unknown"),
                })

        # BUGFIX: previously this ran for ANY export as long as there were
        # visible boxes, so a "_corrected_detection.shp" was created even if
        # the user never eliminated/added/edited a single box (every box
        # defaults to status="retained"). That made "corrected" misleading --
        # it was identical to the raw export except for the status field.
        # Only write it when the undo stack shows an actual correction action
        # happened (eliminate/add/edit), regardless of whether it was later
        # undone back to a state with boxes still remaining.
        corrected_was_generated = bool(corr_boxes and user_made_corrections)
        if corrected_was_generated:
            save_corrected_shapefile(
                raster_path=p_raster,
                boxes=np.array(corr_boxes, dtype=np.float32),
                scores=np.array(corr_scores, dtype=np.float32),
                classes=np.array(corr_classes, dtype=np.int32),
                statuses=corr_statuses,
                out_shp=corrected_shp,
                model_name="yolo_multispectral",
                class_names=class_names_corr,
                validator_name=reviewer_name,
                correction_date=correction_date,
            )
        elif corr_boxes and not user_made_corrections:
            self.logger.info(
                "No manual corrections detected -- skipping "
                f"'{corrected_shp.name}' (identical to raw export otherwise)."
            )

        excel_path = self._export_inference_excel(
            p_raster,
            out_dir,
            output_stem,
            active_box_records,
            fast_mode=True,
        )
        geojson_path = self._export_inference_centroid_geojson(p_raster, out_dir, output_stem, active_box_records)

        # BUGFIX: this list used to hardcode "2. Corrected: ..." regardless
        # of whether save_corrected_shapefile() above actually ran, so the
        # success dialog kept advertising a _corrected_detection.shp file
        # even on exports with no manual corrections (where it's correctly
        # skipped on disk). Only list files that were genuinely written, and
        # number them sequentially so there's no gap.
        summary = [f"Raw: {raw_shp.name}"]
        if corrected_was_generated:
            summary.append(f"Corrected: {corrected_shp.name}")
        if excel_path is not None:
            summary.append(f"Excel: {excel_path.name}")
        if geojson_path is not None:
            summary.append(f"GeoJSON: {geojson_path.name}")
        summary = [f"{i}. {line}" for i, line in enumerate(summary, start=1)]
        QMessageBox.information(
            mw, "Export Successful",
            "Successfully exported inference outputs!\n\n" +
            "\n".join(summary) +
            f"\n\nFolder: {out_dir}"
        )

    def _export_inference_centroid_metrics(self, raster_path: Path, out_dir: Path, output_stem: Optional[str] = None):
        """Export centroid metrics from corrected inference boxes as a shapefile."""
        if not self.box_items:
            return None

        try:
            import shapefile
            from shapely.geometry import Point
            from utils.geospatial_utils import GeospatialMetrics
            import numpy as np

            active_boxes = [item for item in self.box_items if item.status != "eliminated" and item.isVisible()]
            if not active_boxes:
                return None

            stem = Path(output_stem).stem if output_stem else raster_path.stem
            shp_name = f"{stem}_metrics.shp"
            shp_path = out_dir / shp_name

            with shapefile.Writer(str(shp_path), shapeType=shapefile.POINT) as shp:
                shp.field("id", "N", size=10)
                shp.field("lat", "F", decimal=8)
                shp.field("lon", "F", decimal=8)
                shp.field("radius_m", "F", decimal=4)
                shp.field("diameter_m", "F", decimal=4)
                shp.field("area_m2", "F", decimal=4)

                # Build metrics helper using raster CRS/transform
                metrics = None
                try:
                    from rasterio import open as rio_open
                    with rio_open(raster_path) as src:
                        transform = src.transform
                        crs = src.crs
                        metrics = GeospatialMetrics(transform, crs)
                        crs_wkt = src.crs.to_wkt() if src.crs else None
                except Exception as e:
                    self.main_window.logger.warning(f"Could not open raster for metrics export: {e}")
                    metrics = None
                    crs_wkt = None

                for idx, item in enumerate(active_boxes, start=1):
                    x1, y1, x2, y2 = item.get_box_coords()
                    cx = (x1 + x2) / 2.0
                    cy = (y1 + y2) / 2.0
                    lon, lat = (0.0, 0.0)
                    if metrics:
                        try:
                            lon, lat = metrics.pixel_to_latlon(cx, cy)
                        except Exception as e:
                            self.main_window.logger.debug(f"Centroid metric conversion failed: {e}")

                    radius_px = min(abs(x2 - x1), abs(y2 - y1)) / 2.0
                    radius_m = 0.0
                    if metrics and metrics.transform is not None:
                        pixel_width_m = abs(metrics.transform.a)
                        pixel_height_m = abs(metrics.transform.e)
                        pixel_size_m = (pixel_width_m + pixel_height_m) / 2.0
                        radius_m = radius_px * pixel_size_m

                    diameter_m = radius_m * 2.0
                    area_m2 = np.pi * (radius_m ** 2)

                    shp.point(lon, lat)
                    shp.record(
                        idx,
                        round(lat, 8),
                        round(lon, 8),
                        round(radius_m, 4),
                        round(diameter_m, 4),
                        round(area_m2, 4)
                    )

            if crs_wkt:
                with open(shp_path.with_suffix('.prj'), 'w') as prj:
                    prj.write(crs_wkt)

            return shp_path
        except ImportError as e:
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                "Missing library for centroid metrics export. Install pyshp and shapely."
            )
            return None
        except Exception as e:
            self.main_window.logger.error(f"Failed centroid metrics export: {e}", exc_info=True)
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                f"Failed to export centroid metrics: {e}"
            )
            return None

    def _export_inference_excel(self, raster_path: Path, out_dir: Path, output_stem: str,
                                records: List[Dict[str, Any]], fast_mode: bool = True) -> Optional[Path]:
        """Export centroid metrics in the compact format: lat, lon, radius_m, diameter_m, area_m2."""
        if not records:
            return None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                "Missing openpyxl. Install openpyxl to export Excel files."
            )
            return None

        try:
            from rasterio import open as rio_open
            from utils.geospatial_utils import GeospatialMetrics
        except ImportError as e:
            self.main_window.logger.warning(f"Missing raster export dependency: {e}")
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                "Missing rasterio. Install rasterio to export Excel files."
            )
            return None

        excel_path = out_dir / f"{Path(output_stem).stem}.xlsx"
        rows = []

        try:
            with rio_open(raster_path) as src:
                transform = src.transform
                crs = src.crs
                metrics = None
                if transform is not None and crs is not None:
                    metrics = GeospatialMetrics(transform, crs)

                for record in records:
                    x1, y1, x2, y2 = record.get('box', [0.0, 0.0, 0.0, 0.0])
                    cx = float((x1 + x2) / 2.0)
                    cy = float((y1 + y2) / 2.0)
                    lon, lat = 0.0, 0.0
                    if metrics:
                        try:
                            lon, lat = metrics.pixel_to_latlon(cx, cy)
                        except Exception as e:
                            self.main_window.logger.debug(f"Excel export lat/lon conversion failed: {e}")

                    radius_px = min(abs(x2 - x1), abs(y2 - y1)) / 2.0
                    pixel_width_m = abs(float(transform.a)) if transform is not None else 1.0
                    pixel_height_m = abs(float(transform.e)) if transform is not None else 1.0
                    pixel_size_m = (pixel_width_m + pixel_height_m) / 2.0 if pixel_width_m and pixel_height_m else 1.0
                    radius_m = radius_px * pixel_size_m
                    diameter_m = radius_m * 2.0
                    area_m2 = 3.141592653589793 * (radius_m ** 2)

                    rows.append([
                        round(lat, 8),
                        round(lon, 8),
                        round(radius_m, 8),
                        round(diameter_m, 8),
                        round(area_m2, 8),
                    ])

            if not rows:
                return None

            wb = Workbook()
            ws = wb.active
            ws.title = "Inference Results"
            headers = ["Latitude", "Longitude", "radius_m", "diameter_m", "area_m2"]
            ws.append(headers)
            for row in rows:
                ws.append(row)

            header_fill = PatternFill("solid", fgColor="1F4E78")
            coord_fill = PatternFill("solid", fgColor="D9EAF7")
            radius_fill = PatternFill("solid", fgColor="E2F0D9")
            diameter_fill = PatternFill("solid", fgColor="FFF2CC")
            area_fill = PatternFill("solid", fgColor="FCE4D6")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(style="thin", color="D9D9D9")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

            fills = [coord_fill, coord_fill, radius_fill, diameter_fill, area_fill]
            for col_idx, fill in enumerate(fills, start=1):
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.fill = fill

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for column_cells in ws.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = max(length + 2, 14)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if cell.column in [3, 4, 5]:
                        cell.number_format = "0.#######"

            try:
                wb.save(str(excel_path))
            except Exception as excel_err:
                self.main_window.logger.error(f"Failed to save Excel export: {excel_err}", exc_info=True)
                QMessageBox.warning(
                    self.main_window,
                    "Export Failed",
                    f"Failed to save Excel file: {excel_err}"
                )
                return None
            return excel_path
        except Exception as e:
            self.main_window.logger.error(f"Failed Excel export: {e}", exc_info=True)
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                f"Failed to export Excel: {e}"
            )
            return None

    def _export_inference_centroid_geojson(self, raster_path: Path, out_dir: Path, output_stem: str,
                                           records: List[Dict[str, Any]]) -> Optional[Path]:
        """Export centroid records as a GeoJSON point file."""
        if not records:
            return None

        try:
            import json
            from rasterio import open as rio_open
            from utils.geospatial_utils import GeospatialMetrics
        except ImportError as e:
            self.main_window.logger.warning(f"Missing geojson export dependency: {e}")
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                "Missing rasterio or json. Install rasterio to export GeoJSON."
            )
            return None

        geojson_path = out_dir / f"{Path(output_stem).stem}_centroid.geojson"
        features = []

        try:
            with rio_open(raster_path) as src:
                transform = src.transform
                crs = src.crs
                metrics = None
                if transform is not None and crs is not None:
                    metrics = GeospatialMetrics(transform, crs)

                for record in records:
                    x1, y1, x2, y2 = record.get('box', [0.0, 0.0, 0.0, 0.0])
                    cx = float((x1 + x2) / 2.0)
                    cy = float((y1 + y2) / 2.0)
                    lon, lat = 0.0, 0.0
                    if metrics:
                        try:
                            lon, lat = metrics.pixel_to_latlon(cx, cy)
                        except Exception as e:
                            self.main_window.logger.debug(f"GeoJSON export lat/lon conversion failed: {e}")

                    feature = {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [lon, lat]
                        },
                        "properties": {
                            "id": record.get('id', None),
                            "class_name": record.get('class_name', None)
                        }
                    }
                    features.append(feature)

            geojson = {
                "type": "FeatureCollection",
                "features": features
            }
            with open(geojson_path, 'w', encoding='utf-8') as f:
                json.dump(geojson, f, ensure_ascii=False, indent=2)
            return geojson_path
        except Exception as e:
            self.main_window.logger.error(f"Failed GeoJSON export: {e}", exc_info=True)
            QMessageBox.warning(
                self.main_window,
                "Export Failed",
                f"Failed to export GeoJSON: {e}"
            )
            return None
