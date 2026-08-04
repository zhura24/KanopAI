"""
inference_engine.py
Engine deteksi sawit dari raster multispektral (Ultralytics YOLO .pt).
Menggantikan alur ONNX lama (core/detection_worker.py + handlers/detection_handler.py).
Logic murni sama persis dengan inference_core.py -- tidak diubah, cuma dipindah
lokasi supaya konsisten dengan struktur folder core/ yang sudah ada.
"""

import json
from pathlib import Path
from dataclasses import dataclass, field

import cv2
import numpy as np
import rasterio
import rasterio.windows
import rasterio.transform
import torch
from ultralytics import YOLO
import gc  # dipakai buat lepas memori tile-batch tiap iterasi -- penting di device RAM 8GB


# ============================================================
# HASIL
# ============================================================
@dataclass
class InferenceResult:
    boxes: np.ndarray = field(default_factory=lambda: np.zeros((0, 4)))
    scores: np.ndarray = field(default_factory=lambda: np.zeros((0,)))
    classes: np.ndarray = field(default_factory=lambda: np.zeros((0,)))
    class_names: list = None
    class_counts: list = None
    shp_path: Path = None
    preview_path: Path = None
    excel_path: Path = None
    preview_bgr: np.ndarray = None  # composite RGB + kotak, siap ditampilkan di canvas
    elapsed_seconds: float = 0.0  # waktu total proses run(), buat Dashboard Card "Waktu Inference"
    n_tiles: int = 0  # jumlah tile yang diproses, buat Dashboard Card "Jumlah Tile"
    preview_scale: float = 1.0  # skala downsample preview_bgr vs resolusi asli raster;
                                 # WAJIB dipakai saat gambar overlay kotak interaktif lain
    aoi_polygons_px: list = None
    exclude_polygons_px: list = None
    run_id: int = None  # id baris di tabel SQLite 'runs' -- hanya terisi kalau db_path dipakai


class CancelledError(Exception):
    pass


def resolve_class_name(class_id, class_names=None):
    if class_names is None:
        return str(int(class_id))
    if isinstance(class_names, dict):
        if class_id in class_names:
            return str(class_names[class_id])
        if str(class_id) in class_names:
            return str(class_names[str(class_id)])
    elif isinstance(class_names, (list, tuple, np.ndarray)):
        idx = int(class_id)
        if 0 <= idx < len(class_names):
            return str(class_names[idx])
    return str(int(class_id))


def summarize_class_counts(classes, class_names=None):
    values = np.asarray(classes, dtype=np.int32).reshape(-1)
    if len(values) == 0:
        return []
    counts = {}
    for cls_id in values:
        key = int(cls_id)
        counts[key] = counts.get(key, 0) + 1
    return [(resolve_class_name(key, class_names), counts[key]) for key in sorted(counts)]


def _polygon_mask_for_window(polygons, row_start, col_start, height, width):
    """Return a pixel-center mask for pixel-coordinate polygons in a window."""
    if not polygons:
        return np.ones((height, width), dtype=bool)

    rows, cols = np.mgrid[
        row_start:row_start + height,
        col_start:col_start + width,
    ]
    x = cols.astype(np.float64) + 0.5
    y = rows.astype(np.float64) + 0.5
    mask = np.zeros((height, width), dtype=bool)

    for polygon in polygons:
        points = np.asarray(polygon, dtype=np.float64)
        if points.ndim != 2 or points.shape[0] < 3 or points.shape[1] < 2:
            continue

        polygon_mask = np.zeros((height, width), dtype=bool)
        x1 = points[:, 0]
        y1 = points[:, 1]
        x2 = np.roll(x1, -1)
        y2 = np.roll(y1, -1)
        for edge_x1, edge_y1, edge_x2, edge_y2 in zip(x1, y1, x2, y2):
            crosses = (edge_y1 > y) != (edge_y2 > y)
            denominator = edge_y2 - edge_y1
            if denominator == 0:
                continue
            intersections = (edge_x2 - edge_x1) * (y - edge_y1) / denominator + edge_x1
            polygon_mask ^= crosses & (x < intersections)
        mask |= polygon_mask

    return mask


def _read_band_values_for_detection(src, band_index, box, aoi_polygons_px=None, exclude_polygons_px=None):
    """Read a center pixel or polygon-clipped mean for one detection band."""
    x1, y1, x2, y2 = [float(value) for value in box]
    if not aoi_polygons_px and not exclude_polygons_px:
        col = int(round((x1 + x2) / 2.0))
        row = int(round((y1 + y2) / 2.0))
        if 0 <= row < src.height and 0 <= col < src.width:
            data = src.read(band_index, window=((row, row + 1), (col, col + 1)))
            if data.size > 0:
                return float(data[0, 0])
        return None

    col_start = max(0, int(np.floor(x1)))
    row_start = max(0, int(np.floor(y1)))
    col_stop = min(src.width, int(np.ceil(x2)))
    row_stop = min(src.height, int(np.ceil(y2)))
    if col_start >= col_stop or row_start >= row_stop:
        return None

    window = ((row_start, row_stop), (col_start, col_stop))
    data = src.read(band_index, window=window)
    height, width = data.shape
    valid_mask = _polygon_mask_for_window(
        aoi_polygons_px, row_start, col_start, height, width
    ) if aoi_polygons_px else np.ones((height, width), dtype=bool)
    if exclude_polygons_px:
        valid_mask &= ~_polygon_mask_for_window(
            exclude_polygons_px, row_start, col_start, height, width
        )

    values = data[valid_mask]
    values = values[np.isfinite(values)]
    if values.size == 0:
        return None
    return float(np.mean(values))


def _read_center_pixel_for_band(src, band_index, x, y):
    """Fast path for export: read one pixel at the box center instead of full box statistics."""
    col = int(round(float(x)))
    row = int(round(float(y)))
    if not (0 <= row < src.height and 0 <= col < src.width):
        return None
    data = src.read(band_index, window=((row, row + 1), (col, col + 1)))
    if data.size == 0:
        return None
    return float(data[0, 0])


def export_result_excel(raster_path, out_path, boxes, scores, classes, class_names=None,
                        aoi_polygons_px=None, exclude_polygons_px=None, fast_mode=False):
    """Export centroid metrics in a compact worksheet: latitude, longitude, radius_m, diameter_m, area_m2."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None

    try:
        from rasterio import open as rio_open
    except ImportError:
        return None

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    boxes = np.asarray(boxes, dtype=np.float32) if boxes is not None else np.zeros((0, 4), dtype=np.float32)
    scores = np.asarray(scores, dtype=np.float32) if scores is not None else np.zeros((0,), dtype=np.float32)
    classes = np.asarray(classes, dtype=np.int32) if classes is not None else np.zeros((0,), dtype=np.int32)

    if len(boxes) == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inference Results"
        ws.append(["Latitude", "Longitude", "radius_m", "diameter_m", "area_m2"])
        wb.save(str(out_path))
        return out_path

    rows = []
    try:
        with rio_open(raster_path) as src:
            transform = src.transform
            crs = src.crs
            metrics = None
            if transform is not None and crs is not None:
                try:
                    from utils.geospatial_utils import GeospatialMetrics
                    metrics = GeospatialMetrics(transform, crs)
                except Exception:
                    metrics = None

            for idx, box in enumerate(boxes, start=1):
                x1, y1, x2, y2 = [float(v) for v in box]
                cx = float((x1 + x2) / 2.0)
                cy = float((y1 + y2) / 2.0)
                lon, lat = 0.0, 0.0
                if metrics is not None:
                    try:
                        lon, lat = metrics.pixel_to_latlon(cx, cy)
                    except Exception:
                        lon, lat = 0.0, 0.0

                radius_px = min(abs(x2 - x1), abs(y2 - y1)) / 2.0
                if transform is not None:
                    pixel_size_m = (abs(float(transform.a)) + abs(float(transform.e))) / 2.0 if transform.a != 0 or transform.e != 0 else 1.0
                else:
                    pixel_size_m = 1.0
                radius_m = radius_px * pixel_size_m
                diameter_m = radius_m * 2.0
                area_m2 = np.pi * (radius_m ** 2)
                rows.append([round(lat, 8), round(lon, 8), round(radius_m, 8), round(diameter_m, 8), round(area_m2, 8)])

        if not rows:
            wb = Workbook()
            ws = wb.active
            ws.title = "Inference Results"
            ws.append(["Latitude", "Longitude", "radius_m", "diameter_m", "area_m2"])
            wb.save(str(out_path))
            return out_path

        wb = Workbook()
        ws = wb.active
        ws.title = "Inference Results"
        ws.append(["Latitude", "Longitude", "radius_m", "diameter_m", "area_m2"])
        for row in rows:
            ws.append(row)
        wb.save(str(out_path))
        return out_path
    except Exception:
        return None


# ============================================================
# FUNGIA MURNI (identik dengan v2 script, dipisah biar mudah ditest)
# ============================================================
def load_band_stats(stats_path: Path) -> dict:
    with open(stats_path, "r") as f:
        raw = json.load(f)
    return {int(k): v for k, v in raw.items()}


def is_multiref_schema(band_stats: dict) -> bool:
    return any("sources" in v for v in band_stats.values())


def stretch_band(band: np.ndarray, p_low: float, p_high: float) -> np.ndarray:
    band = band.astype(np.float32)  # float32 cukup (raster & hasil akhir uint8), hemat RAM 2x vs float64
    if p_high - p_low == 0:
        return np.zeros_like(band, dtype=np.uint8)
    clipped = np.clip(band, p_low, p_high)
    scaled = (clipped - p_low) / (p_high - p_low) * 255.0
    return scaled.astype(np.uint8)


def generate_tile_windows(width, height, tile_size=640, overlap=64):
    stride = tile_size - overlap
    windows = []
    y = 0
    while y < height:
        x = 0
        h = min(tile_size, height - y)
        while x < width:
            w = min(tile_size, width - x)
            windows.append((x, y, w, h))
            if x + tile_size >= width:
                break
            x += stride
        if y + tile_size >= height:
            break
        y += stride
    return windows


def pad_tile_for_inference(tile_hwc: np.ndarray, target_size: int = 640) -> np.ndarray:
    height, width = tile_hwc.shape[:2]
    if height >= target_size and width >= target_size:
        return tile_hwc
    pad_h = max(0, target_size - height)
    pad_w = max(0, target_size - width)
    if pad_h == 0 and pad_w == 0:
        return tile_hwc
    return np.pad(tile_hwc, ((0, pad_h), (0, pad_w), (0, 0)), mode="constant", constant_values=0)


# ============================================================
# DEDUPLIKASI TILE-BOUNDARY -- PENDEKATAN "KEPEMILIKAN WILAYAH"
# ============================================================
def compute_tile_core_bounds(x_off, y_off, w, h, width, height, overlap):
    """
    Hitung wilayah "inti" (core) satu tile -- area yang jadi tanggung jawab
    TUNGGAL tile ini untuk mendeteksi objek, tanpa berbagi dengan tile
    tetangga. Ini pengganti pendekatan lama (merge berbasis jarak/IoU
    sesudah inference) dengan cara yang jauh lebih pasti: alih-alih menebak
    "dua box ini duplikat atau bukan" dari kemiripan geometrinya (yang bisa
    salah kalau kanopi kecil & rapat, lihat histori chat), setiap titik di
    raster SEJAK AWAL cuma "dimiliki" oleh SATU tile saja -- ditentukan murni
    dari posisi tile itu sendiri, sama sekali tidak bergantung pada isi
    deteksi. Jadi tidak ada lagi ambiguitas "pohon sama vs pohon beda".

    Aturan: potong separuh lebar overlap dari sisi yang berbatasan dengan
    tile tetangga (kiri/kanan/atas/bawah), TAPI kalau sisi itu adalah tepi
    raster asli (tidak ada tetangga di situ), tidak dipotong sama sekali --
    supaya tidak ada gap yang kehilangan cakupan di pinggir gambar.

    Karena generate_tile_windows() selalu memberi overlap yang KONSTAN antar
    tile bertetangga (termasuk tile terakhir di tiap baris/kolom, karena tile
    terakhir selalu digeser pas rata ke tepi raster), potongan overlap/2 yang
    tetap ini otomatis pas untuk semua pasangan tile tanpa perlu kalkulasi
    khusus per tile.
    """
    half = overlap // 2
    core_x_min = x_off if x_off == 0 else x_off + half
    core_x_max = (x_off + w) if (x_off + w) >= width else (x_off + w - half)
    core_y_min = y_off if y_off == 0 else y_off + half
    core_y_max = (y_off + h) if (y_off + h) >= height else (y_off + h - half)
    return core_x_min, core_x_max, core_y_min, core_y_max


def filter_by_tile_ownership(boxes: np.ndarray, tile_ids: np.ndarray,
                              windows_by_id: dict, width: int, height: int,
                              overlap: int) -> np.ndarray:
    """
    Buang deteksi yang TITIK TENGAHNYA jatuh di luar wilayah inti tile asalnya
    (lihat compute_tile_core_bounds). Untuk objek yang muncul di zona overlap
    dan terdeteksi dari 2 tile berbeda, hanya salinan dari tile yang memang
    "memiliki" lokasi itu yang dipertahankan -- salinan dari tile lain otomatis
    tersingkir, TANPA perlu menebak dari jarak/ukuran/IoU sama sekali.

    Return: boolean mask (True = pertahankan).
    """
    if len(boxes) == 0:
        return np.zeros(0, dtype=bool)

    max_id = max(tile_ids.max(), max(windows_by_id.keys()))
    max_id = int(max_id)
    cxmin_lut = np.zeros(max_id + 1)
    cxmax_lut = np.zeros(max_id + 1)
    cymin_lut = np.zeros(max_id + 1)
    cymax_lut = np.zeros(max_id + 1)
    for tid, (x_off, y_off, w, h) in windows_by_id.items():
        bounds = compute_tile_core_bounds(x_off, y_off, w, h, width, height, overlap)
        cxmin_lut[tid], cxmax_lut[tid], cymin_lut[tid], cymax_lut[tid] = bounds

    cx = (boxes[:, 0] + boxes[:, 2]) / 2.0
    cy = (boxes[:, 1] + boxes[:, 3]) / 2.0

    core_x_min = cxmin_lut[tile_ids]
    core_x_max = cxmax_lut[tile_ids]
    core_y_min = cymin_lut[tile_ids]
    core_y_max = cymax_lut[tile_ids]

    keep = (cx >= core_x_min) & (cx < core_x_max) & (cy >= core_y_min) & (cy < core_y_max)
    return keep


def nms_global(boxes: np.ndarray, scores: np.ndarray, classes: np.ndarray = None, iou_threshold: float = 0.5):
    """
    NMS standar berbasis IoU -- membuang box yang tumpang tindih tinggi
    dengan box skor lebih tinggi. Dilakukan per-kelas jika 'classes' diberikan.
    """
    if len(boxes) == 0:
        return []
        
    if classes is not None and len(classes) > 0:
        max_coordinate = boxes.max() if boxes.size > 0 else 0
        offsets = classes * (max_coordinate + 1)
        _boxes = boxes + offsets[:, None]
    else:
        _boxes = boxes

    x1, y1, x2, y2 = _boxes[:, 0], _boxes[:, 1], _boxes[:, 2], _boxes[:, 3]
    areas = (x2 - x1) * (y2 - y1)

    order = scores.argsort()[::-1]
    keep = []
    while order.size > 0:
        i = order[0]
        keep.append(i)

        rest = order[1:]

        xx1 = np.maximum(x1[i], x1[rest])
        yy1 = np.maximum(y1[i], y1[rest])
        xx2 = np.minimum(x2[i], x2[rest])
        yy2 = np.minimum(y2[i], y2[rest])
        w = np.maximum(0.0, xx2 - xx1)
        h = np.maximum(0.0, yy2 - yy1)
        inter = w * h
        iou = inter / (areas[i] + areas[rest] - inter)

        is_duplicate = iou > iou_threshold
        order = rest[~is_duplicate]
    return keep


# ============================================================
# AOI / EXCLUDE ZONE -- FITUR OPSIONAL
# ============================================================
# Sengaja TIDAK pakai shapely/geopandas (dependency itu sudah dilepas dari
# proyek ini). Polygon AOI/exclude dibaca lewat pyshp (sudah dipakai untuk
# shapefile deteksi), lalu point-in-polygon dicek manual pakai ray-casting.
# Koordinat polygon dikonversi dari geo -> piksel raster memakai transform
# terbalik, supaya bisa langsung dibandingkan dengan koordinat box (x1_px,
# y1_px, x2_px, y2_px) yang sudah dipakai di seluruh pipeline ini.

def load_polygons_px(shp_path: Path, raster_transform) -> list:
    """Baca semua polygon dari sebuah shapefile (AOI atau exclude area) dan
    kembalikan sebagai list array Nx2 dalam koordinat PIKSEL raster."""
    import shapefile  # pyshp

    polys = []
    with shapefile.Reader(str(shp_path)) as shp:
        for rec_shape in shp.shapes():
            geo_pts = rec_shape.points
            if len(geo_pts) < 3:
                continue
            px_pts = []
            for gx, gy in geo_pts:
                col, row = ~raster_transform * (gx, gy)
                px_pts.append((col, row))
            polys.append(np.asarray(px_pts, dtype=np.float64))
    return polys


def _point_in_polygon(x: float, y: float, poly: np.ndarray) -> bool:
    """Ray-casting standar, tanpa dependency tambahan."""
    n = len(poly)
    inside = False
    j = n - 1
    for i in range(n):
        xi, yi = poly[i]
        xj, yj = poly[j]
        if (yi > y) != (yj > y):
            x_intersect = (xj - xi) * (y - yi) / (yj - yi + 1e-12) + xi
            if x < x_intersect:
                inside = not inside
        j = i
    return inside


def filter_boxes_by_aoi(boxes: np.ndarray, aoi_polys: list = None,
                         exclude_polys: list = None) -> np.ndarray:
    """Buang box yang TITIK TENGAHNYA jatuh di luar semua polygon AOI
    (kalau AOI diisi), atau jatuh DI DALAM salah satu polygon exclude area.
    Kalau aoi_polys dan exclude_polys dua-duanya kosong/None, semua box
    dipertahankan (fitur ini murni opsional, tidak mengubah perilaku lama
    kalau tidak dipakai)."""
    n = len(boxes)
    if not aoi_polys and not exclude_polys:
        return np.ones(n, dtype=bool)

    cx = (boxes[:, 0] + boxes[:, 2]) / 2.0
    cy = (boxes[:, 1] + boxes[:, 3]) / 2.0
    keep = np.ones(n, dtype=bool)

    if aoi_polys:
        in_aoi = np.zeros(n, dtype=bool)
        for i in range(n):
            for poly in aoi_polys:
                if _point_in_polygon(cx[i], cy[i], poly):
                    in_aoi[i] = True
                    break
        keep &= in_aoi

    if exclude_polys:
        for i in range(n):
            if not keep[i]:
                continue
            for poly in exclude_polys:
                if _point_in_polygon(cx[i], cy[i], poly):
                    keep[i] = False
                    break

    return keep


def tile_intersects_aoi(x_off, y_off, w, h, aoi_polys: list) -> bool:
    """Cek cepat berbasis bounding-box (bukan cek polygon presisi) untuk
    SKIP tile yang jelas-jelas di luar AOI sebelum tile itu dibaca & di-
    infer -- ini yang paling membantu untuk device RAM 8GB, karena tile
    yang jelas tidak relevan tidak usah dialokasikan sama sekali."""
    if not aoi_polys:
        return True
    tile_x_min, tile_x_max = x_off, x_off + w
    tile_y_min, tile_y_max = y_off, y_off + h
    for poly in aoi_polys:
        px_min, py_min = poly[:, 0].min(), poly[:, 1].min()
        px_max, py_max = poly[:, 0].max(), poly[:, 1].max()
        if tile_x_max >= px_min and tile_x_min <= px_max and tile_y_max >= py_min and tile_y_min <= py_max:
            return True
    return False


def polygons_from_pixel_coords(polygon_coords_list: list) -> list:
    """Konversi list koordinat piksel [[(x,y),...], ...] ke list array Nx2."""
    polys = []
    for coords in polygon_coords_list or []:
        if not coords or len(coords) < 3:
            continue
        polys.append(np.asarray(coords, dtype=np.float64))
    return polys


def compute_visible_crop_region(width: int, height: int, viewport_rect: tuple) -> tuple:
    """Clamp viewport bounds to raster dimensions and return (x, y, w, h)."""
    x0, y0, x1, y1 = viewport_rect
    x_min = max(0, min(int(x0), width))
    y_min = max(0, min(int(y0), height))
    x_max = min(width, max(int(x1), x_min))
    y_max = min(height, max(int(y1), y_min))
    if x_max <= x_min or y_max <= y_min:
        return 0, 0, width, height
    return x_min, y_min, max(1, x_max - x_min), max(1, y_max - y_min)


def compute_aoi_crop_region(aoi_polys: list, width: int, height: int) -> tuple:
    """Hitung region crop (x, y, w, h) dari union bounding-box polygon AOI."""
    if not aoi_polys:
        return 0, 0, width, height
    xs = np.concatenate([p[:, 0] for p in aoi_polys])
    ys = np.concatenate([p[:, 1] for p in aoi_polys])
    x_min = max(0, int(np.floor(xs.min())))
    y_min = max(0, int(np.floor(ys.min())))
    x_max = min(width, int(np.ceil(xs.max())))
    y_max = min(height, int(np.ceil(ys.max())))
    crop_w = max(1, x_max - x_min)
    crop_h = max(1, y_max - y_min)
    return x_min, y_min, crop_w, crop_h


def rasterize_polygon_mask(poly_px: np.ndarray, x_off: int, y_off: int,
                           w: int, h: int) -> np.ndarray:
    """Rasterize satu polygon (koordinat piksel global) ke mask boolean tile."""
    rel = poly_px.copy()
    rel[:, 0] -= x_off
    rel[:, 1] -= y_off
    pts = rel.astype(np.int32).reshape((-1, 1, 2))
    mask = np.zeros((h, w), dtype=np.uint8)
    cv2.fillPoly(mask, [pts], 1)
    return mask.astype(bool)


def apply_polygon_masks_to_tile(tile_chw: np.ndarray, x_off: int, y_off: int,
                                w: int, h: int, aoi_polys: list = None,
                                exclude_polys: list = None) -> np.ndarray:
    """Mask tile: nol di luar AOI dan di dalam exclude area."""
    if not aoi_polys and not exclude_polys:
        return tile_chw

    keep = np.ones((h, w), dtype=bool)
    if aoi_polys:
        aoi_mask = np.zeros((h, w), dtype=bool)
        for poly in aoi_polys:
            aoi_mask |= rasterize_polygon_mask(poly, x_off, y_off, w, h)
        keep &= aoi_mask

    if exclude_polys:
        for poly in exclude_polys:
            keep &= ~rasterize_polygon_mask(poly, x_off, y_off, w, h)

    if not keep.all():
        tile_chw = tile_chw.copy()
        tile_chw[:, ~keep] = 0
    return tile_chw


def auto_detect_band_mapping_multiref(src, band_stats: dict, log=print) -> dict:
    """Greedy multireference band matching, optimized for large rasters.

    Root cause of the slowness: the previous implementation recomputed the full
    remaining candidate matrix on every assignment, effectively doing repeated
    O(S * C * I) work while also re-scanning all remaining slots/sources. For
    large rasters and multireference sensor definitions, that can become
    surprisingly expensive even though the algorithm is only "greedy".

    The optimized version keeps the exact same selection strategy, but it first
    groups candidates by slot and then for each candidate picks only the best
    remaining input band instead of comparing every remaining slot/source against
    every remaining input band again.
    """
    n_bands_input = src.count
    input_means = {}
    for b in range(1, n_bands_input + 1):
        # Downsample read to avoid OOM on large images (e.g. 15423 x 20056)
        h_new = max(1, src.height // 10)
        w_new = max(1, src.width // 10)
        data = src.read(b, out_shape=(h_new, w_new)).astype(np.float32)
        valid = data[data > 0]
        input_means[b] = float(valid.mean()) if len(valid) > 0 else 0.0

    candidates_by_slot = {}
    for slot, entry in band_stats.items():
        slot_candidates = []
        for source_name, stats in entry["sources"].items():
            source_mean = float(stats.get("mean", 0.0))
            slot_candidates.append((source_name, source_mean, stats))
        candidates_by_slot[slot] = slot_candidates

    available_slots = list(band_stats.keys())
    available_input = list(input_means.keys())
    mapping = {}

    while available_slots and available_input:
        best = None
        best_cost = float("inf")

        for slot in available_slots:
            for source_name, mean_val, stats in candidates_by_slot.get(slot, []):
                best_ib = None
                best_ib_diff = None
                for ib in available_input:
                    diff = abs(mean_val - input_means[ib])
                    if best_ib is None or diff < best_ib_diff:
                        best_ib = ib
                        best_ib_diff = diff

                if best_ib is None:
                    continue

                if best_ib_diff < best_cost:
                    best_cost = best_ib_diff
                    best = (best_ib_diff, slot, best_ib, source_name, stats)

        if best is None:
            break

        diff, slot, ib, source_name, stats = best
        mapping[slot] = {
            "input_band": ib,
            "source": source_name,
            "p_low": stats.get("p_low"),
            "p_high": stats.get("p_high"),
        }
        available_slots.remove(slot)
        available_input.remove(ib)
        flag = "  <-- selisih besar, VERIFIKASI MANUAL" if diff > 0.05 else ""
        log(f"  -> Slot {slot} <- band input {ib} (sumber: {source_name}, diff={diff:.6f}){flag}")

    return mapping


def compute_adaptive_norm_params(src, band_idx: int, p_low_pct: float, p_high_pct: float) -> tuple:
    """KF-03 mode 3 (fallback adaptif) -- hitung parameter normalisasi
    (persentil p_low/p_high) LANGSUNG dari data raster input band ke-`band_idx`
    itu sendiri, BUKAN dipinjam dari band_stats sensor lain yang tidak relevan.

    Dipakai untuk sensor yang belum dikenali sistem (band count/karakteristik
    beda dari semua referensi di band_stats). Baca versi downsample dulu
    (sama seperti auto_detect_band_mapping*) supaya tidak OOM di raster besar.
    """
    h_new = max(1, src.height // 10)
    w_new = max(1, src.width // 10)
    data = src.read(band_idx, out_shape=(h_new, w_new)).astype(np.float32)
    valid = data[data > 0]
    if len(valid) == 0:
        return 0.0, 255.0
    p_low, p_high = np.percentile(valid, (p_low_pct, p_high_pct))
    if p_high <= p_low:
        p_low, p_high = float(valid.min()), float(valid.max() + 1e-6)
    return float(p_low), float(p_high)


def auto_detect_band_mapping_adaptive(src, band_stats: dict, p_low_pct: float,
                                       p_high_pct: float, log=print) -> dict:
    """KF-03 mode 3 -- dipakai HANYA saat sensor tidak dikenali sistem
    (jumlah/karakteristik band berbeda dari semua referensi yang ada) DAN
    pengguna sengaja mengaktifkan enable_adaptive_fallback=True di run().

    Urutan band input dipetakan langsung ke urutan slot training (band ke-i
    input -> slot ke-i training), TANPA menebak berdasarkan kemiripan mean
    ke band_stats sensor lain (yang tidak relevan untuk sensor baru). Yang
    dihitung otomatis dari data raster yang sedang diproses hanyalah
    parameter normalisasinya (p_low/p_high per band), lewat
    compute_adaptive_norm_params -- bukan mapping bandnya.

    Ini jaring pengaman sementara, bukan pengganti fine-tuning model untuk
    sensor baru tersebut.
    """
    training_slots = sorted(band_stats.keys())
    n_bands_input = src.count
    mapping = {}
    for i, slot in enumerate(training_slots, start=1):
        if i > n_bands_input:
            log(f"  [PERINGATAN] Slot {slot} tidak ada padanan band input "
                f"(raster cuma {n_bands_input} band) -- dilewati.")
            continue
        p_low, p_high = compute_adaptive_norm_params(src, i, p_low_pct, p_high_pct)
        mapping[slot] = {
            "input_band": i,
            "source": "adaptive_fallback",
            "p_low": p_low,
            "p_high": p_high,
        }
        log(f"  -> Slot {slot} <- band input {i} (fallback adaptif, "
            f"p_low={p_low:.2f}, p_high={p_high:.2f}, dihitung dari data raster ini)")
    return mapping


def validate_manual_band_mapping(manual_mapping: dict, band_stats: dict,
                                  n_bands_input: int, log=print) -> dict:
    """MATCHING BAND MANUAL -- FITUR OPSIONAL.

    Dipakai kalau pengguna tidak percaya/tidak mau pakai auto-detect (mis.
    band mean-nya kebetulan mirip padahal salah slot, atau untuk keperluan
    verifikasi manual). GUI cukup kirim dict {slot_training: band_input},
    fungsi ini yang validasi & normalisasi ke format siap pakai _prepare_tile
    (format sama persis dengan output auto_detect_band_mapping*, jadi tidak
    perlu ubah apa pun di logic pembacaan tile).

    manual_mapping: dict {target_band(int): input_band(int)}.
    Raise ValueError kalau ada slot training yang belum diisi pengguna, atau
    band input yang ditunjuk di luar rentang band raster.
    """
    expected_slots = set(band_stats.keys())
    given_slots = set(int(k) for k in manual_mapping.keys())

    missing = expected_slots - given_slots
    if missing:
        raise ValueError(
            f"Mapping manual belum lengkap -- slot band training berikut "
            f"belum diisi: {sorted(missing)}"
        )

    extra = given_slots - expected_slots
    if extra:
        log(f"[PERINGATAN] Slot {sorted(extra)} di mapping manual tidak "
            f"dikenali band_stats, diabaikan.")

    multiref = is_multiref_schema(band_stats)
    normalized = {}
    used_inputs = {}

    for slot in expected_slots:
        input_b = int(manual_mapping[slot])
        if not (1 <= input_b <= n_bands_input):
            raise ValueError(
                f"Slot {slot}: band input {input_b} di luar rentang "
                f"(raster ini cuma punya {n_bands_input} band)."
            )
        used_inputs.setdefault(input_b, []).append(slot)

        if multiref:
            # fallback p_low/p_high dari source pertama slot ini -- fallback
            # HANYA dipakai kalau tile terkait sangat sedikit piksel valid
            # (<5%), di luar itu tetap distretch dari persentil tile sendiri
            entry = band_stats[slot]
            first_source = next(iter(entry["sources"].values()))
            normalized[slot] = {
                "input_band": input_b,
                "source": "manual",
                "p_low": first_source.get("p_low"),
                "p_high": first_source.get("p_high"),
            }
        else:
            # skema lama: cukup int, _prepare_tile ambil fallback p_low/p_high
            # langsung dari band_stats[slot]
            normalized[slot] = input_b

        log(f"  -> Slot {slot} <- band input {input_b} (manual, dari pengguna)")

    dup = {b: slots for b, slots in used_inputs.items() if len(slots) > 1}
    if dup:
        log(f"[PERINGATAN] Band input berikut dipakai untuk lebih dari 1 slot: "
            f"{dup}. Pastikan ini memang disengaja, bukan salah pilih di GUI.")

    return normalized


def build_preview_bgr(raster_path: Path, boxes: np.ndarray, scores: np.ndarray,
                       stretch_lower_pct: float, stretch_upper_pct: float,
                       max_dim: int = 2000, classes: np.ndarray = None):
    """Composite RGB (band 1-3) dari raster asli + kotak deteksi.
    Return (rgb_bgr, scale) -- 'scale' WAJIB dipakai oleh pemanggil untuk
    menskalakan overlay kotak interaktif lain (mis. QGraphicsRectItem di
    main_window.py), karena gambar yang dikembalikan di sini adalah versi
    DOWNSAMPLE dari raster asli (bukan resolusi penuh) ketika raster lebih
    besar dari max_dim. Tanpa scale ini, overlay lain yang masih pakai
    koordinat piksel resolusi-penuh akan tergambar jauh lebih besar dan
    meleset keluar dari batas gambar preview."""
    with rasterio.open(raster_path) as src:
        h_orig, w_orig = src.height, src.width
        scale = 1.0
        if max(h_orig, w_orig) > max_dim:
            scale = max_dim / max(h_orig, w_orig)
        h_new = int(h_orig * scale)
        w_new = int(w_orig * scale)

        n_bands = src.count
        idx_r = min(3, n_bands)
        idx_g = min(2, n_bands)
        idx_b = min(1, n_bands)

        # Read downsampled version to avoid huge memory footprint
        r = src.read(idx_r, out_shape=(h_new, w_new)).astype(np.float32)
        g = src.read(idx_g, out_shape=(h_new, w_new)).astype(np.float32)
        b = src.read(idx_b, out_shape=(h_new, w_new)).astype(np.float32)

    def stretch_for_display(band):
        valid_pixels = band[band > 0]
        if len(valid_pixels) == 0:
            return np.zeros_like(band, dtype=np.uint8)
        p_low, p_high = np.percentile(valid_pixels, (stretch_lower_pct, stretch_upper_pct))
        if p_high - p_low == 0:
            return np.zeros_like(band, dtype=np.uint8)
        band = np.clip(band, p_low, p_high)
        return ((band - p_low) / (p_high - p_low) * 255).astype(np.uint8)

    rgb = np.stack([stretch_for_display(r), stretch_for_display(g), stretch_for_display(b)], axis=-1)
    rgb_bgr = cv2.cvtColor(rgb, cv2.COLOR_RGB2BGR)

    class_colors = [(0, 255, 0), (0, 165, 255), (255, 0, 0), (255, 255, 0), (255, 0, 255)]
    for idx, (box, score) in enumerate(zip(boxes, scores)):
        # Scale bounding box coordinates to match downsampled preview image
        x1, y1, x2, y2 = (box * scale).astype(int)
        cls_id = int(classes[idx]) if classes is not None and len(classes) > idx else 0
        color = class_colors[cls_id % len(class_colors)]
        cv2.rectangle(rgb_bgr, (x1, y1), (x2, y2), color, 2)
        # Catatan: label confidence sengaja tidak digambar di preview ini
        # (biar lebih bersih dilihat untuk raster dengan ribuan deteksi).
        # Confidence tetap tersedia lewat klik kotak di canvas interaktif,
        # serta tetap tercatat penuh di shapefile/CSV/Excel export.

    return rgb_bgr, scale


def save_shapefile(raster_path: Path, boxes, scores, classes, out_shp: Path, model_name: str = "model_gabungan", class_names=None):
    import shapefile  # pyshp

    with rasterio.open(raster_path) as src:
        raster_transform = src.transform
        if src.crs is not None:
            crs_wkt = src.crs.to_wkt()
            _crs_is_fallback = False
        else:
            # FALLBACK DARURAT: raster tidak memiliki CRS.
            # File .prj ditulis dengan EPSG:4326 agar shapefile tetap bisa
            # dibuka di QGIS/ArcGIS -- TAPI posisi spasialnya TIDAK BISA
            # DIPERCAYA. Koordinat yang ditulis ke shapefile ini berasal
            # dari transform raster yang tidak diketahui artinya (mungkin
            # hanya koordinat piksel, mungkin meter lokal, bukan lat/lon).
            # SOLUSI YANG BENAR: lakukan georeferencing raster input terlebih
            # dahulu menggunakan QGIS (Layer > Georeferencer) atau gdalwarp
            # sebelum menjalankan inference.
            from rasterio.crs import CRS as _RasterioCRS
            crs_wkt = _RasterioCRS.from_epsg(4326).to_wkt()
            _crs_is_fallback = True

    with shapefile.Writer(str(out_shp), shapeType=shapefile.POLYGON) as shp:
        shp.field("id", "N", size=10)
        shp.field("kelas", "C", size=20)
        shp.field("confidence", "N", size=10, decimal=4)
        shp.field("model", "C", size=30)
        shp.field("x1_px", "N", size=10, decimal=1)
        shp.field("y1_px", "N", size=10, decimal=1)
        shp.field("x2_px", "N", size=10, decimal=1)
        shp.field("y2_px", "N", size=10, decimal=1)
        # Field status -- shapefile hasil run() ini adalah "mentah" (belum
        # dikoreksi manusia). Nilainya selalu "not_corrected" di sini;
        # shapefile KE-2 (hasil koreksi pengguna) ditulis terpisah lewat
        # save_corrected_shapefile(), lihat di bawah.
        shp.field("status", "C", size=20)

        for i, (cls, score, box) in enumerate(zip(classes, scores, boxes), start=1):
            x1_px, y1_px, x2_px, y2_px = box
            x1_geo, y1_geo = rasterio.transform.xy(raster_transform, y1_px, x1_px)
            x2_geo, y2_geo = rasterio.transform.xy(raster_transform, y2_px, x2_px)
            polygon = [[x1_geo, y1_geo], [x2_geo, y1_geo], [x2_geo, y2_geo], [x1_geo, y2_geo], [x1_geo, y1_geo]]
            class_name = resolve_class_name(int(cls), class_names)
            shp.poly([polygon])
            shp.record(i, class_name, round(float(score), 4), model_name,
                       round(float(x1_px), 1), round(float(y1_px), 1),
                       round(float(x2_px), 1), round(float(y2_px), 1),
                       "not_corrected")

    # KNF-04: CRS harus konsisten dengan raster asal agar shapefile bisa
    # langsung dibuka di QGIS/software GIS lain.
    if crs_wkt:
        with open(out_shp.with_suffix(".prj"), "w") as prj:
            prj.write(crs_wkt)
    if _crs_is_fallback:
        import logging as _logging
        _logging.getLogger(__name__).warning(
            f"[PERINGATAN CRS] Raster '{raster_path.name}' tidak memiliki CRS. "
            f"Shapefile '{out_shp.name}' ditulis dengan fallback EPSG:4326 agar "
            f"bisa dibuka di GIS, TAPI posisi koordinatnya salah/tidak dapat "
            f"dipercaya. Lakukan georeferencing pada raster input terlebih dahulu."
        )


def save_corrected_shapefile(raster_path: Path, boxes, scores, classes, statuses,
                              out_shp: Path, model_name: str = "model_gabungan",
                              class_names=None, validator_name: str = None,
                              correction_date: str = None):
    """KF-10 -- Shapefile KE-2, ditulis SETELAH pengguna mengoreksi hasil
    deteksi mentah dari save_shapefile() di GUI (mengeliminasi kotak yang
    merupakan false positive/gagal deteksi; TIDAK menambah kotak baru --
    fitur penambahan kotak manual sudah terpisah di aplikasi).

    statuses: array/list sepanjang boxes, with values such as "retained" or
    "eliminated". Semua kotak hasil deteksi awal tetap ditulis di sini
    (bukan cuma yang dipertahankan) untuk
    menjaga audit trail lengkap sesuai KNF-09.
    """
    import shapefile  # pyshp
    from datetime import datetime

    with rasterio.open(raster_path) as src:
        raster_transform = src.transform
        if src.crs is not None:
            crs_wkt = src.crs.to_wkt()
            _crs_is_fallback = False
        else:
            # FALLBACK DARURAT: lihat penjelasan lengkap di save_shapefile().
            from rasterio.crs import CRS as _RasterioCRS
            crs_wkt = _RasterioCRS.from_epsg(4326).to_wkt()
            _crs_is_fallback = True

    with shapefile.Writer(str(out_shp), shapeType=shapefile.POLYGON) as shp:
        shp.field("id", "N", size=10)
        shp.field("kelas", "C", size=20)
        shp.field("confidence", "N", size=10, decimal=4)
        shp.field("model", "C", size=30)
        shp.field("x1_px", "N", size=10, decimal=1)
        shp.field("y1_px", "N", size=10, decimal=1)
        shp.field("x2_px", "N", size=10, decimal=1)
        shp.field("y2_px", "N", size=10, decimal=1)
        shp.field("status", "C", size=20)
        shp.field("pengoreksi", "C", size=50)
        # DBF field names are limited to 10 characters.
        shp.field("tgl_koreks", "C", size=20)

        tgl = correction_date or datetime.now().strftime("%Y-%m-%d")
        for i, (cls, score, box, status) in enumerate(zip(classes, scores, boxes, statuses), start=1):
            x1_px, y1_px, x2_px, y2_px = box
            x1_geo, y1_geo = rasterio.transform.xy(raster_transform, y1_px, x1_px)
            x2_geo, y2_geo = rasterio.transform.xy(raster_transform, y2_px, x2_px)
            polygon = [[x1_geo, y1_geo], [x2_geo, y1_geo], [x2_geo, y2_geo], [x1_geo, y2_geo], [x1_geo, y1_geo]]
            class_name = resolve_class_name(int(cls), class_names)
            shp.poly([polygon])
            shp.record(i, class_name, round(float(score), 4), model_name,
                       round(float(x1_px), 1), round(float(y1_px), 1),
                       round(float(x2_px), 1), round(float(y2_px), 1),
                       str(status), validator_name or "", tgl)

    if crs_wkt:
        with open(out_shp.with_suffix(".prj"), "w") as prj:
            prj.write(crs_wkt)
    if _crs_is_fallback:
        import logging as _logging
        _logging.getLogger(__name__).warning(
            f"[PERINGATAN CRS] Raster '{raster_path.name}' tidak memiliki CRS. "
            f"Shapefile '{out_shp.name}' ditulis dengan fallback EPSG:4326 agar "
            f"bisa dibuka di GIS, TAPI posisi koordinatnya salah/tidak dapat "
            f"dipercaya. Lakukan georeferencing pada raster input terlebih dahulu."
        )


def load_detection_from_shapefile(shp_path: Path):
    """Baca hasil deteksi lama dari shapefile yang dihasilkan aplikasi."""
    import shapefile

    shp_path = Path(shp_path)
    if not shp_path.is_file():
        raise FileNotFoundError(f"Shapefile tidak ditemukan: {shp_path}")

    with shapefile.Reader(str(shp_path)) as shp:
        fields = [f[0] for f in shp.fields[1:]]
        boxes = []
        scores = []
        class_labels = []

        for record in shp.iterRecords():
            values = dict(zip(fields, record))
            x1 = float(values.get("x1_px", 0.0))
            y1 = float(values.get("y1_px", 0.0))
            x2 = float(values.get("x2_px", 0.0))
            y2 = float(values.get("y2_px", 0.0))
            boxes.append([x1, y1, x2, y2])
            scores.append(float(values.get("confidence", 0.0)))
            class_labels.append(str(values.get("kelas", "")) or "sawit")

    if not boxes:
        return (
            np.zeros((0, 4), dtype=np.float32),
            np.zeros((0,), dtype=np.float32),
            np.zeros((0,), dtype=np.float32),
            [],
        )

    unique_names = []
    for label in class_labels:
        if label not in unique_names:
            unique_names.append(label)
    classes = np.array([unique_names.index(label) for label in class_labels], dtype=np.float32)

    return (
        np.asarray(boxes, dtype=np.float32),
        np.asarray(scores, dtype=np.float32),
        classes,
        unique_names,
    )


def _derive_box_from_shape(shape, raster_transform=None, raster_bounds=None):
    """Return pixel-space box from the actual polygon geometry when available.

    The real geometry is the source of truth. DBF fields like x1_px/y1_px can be
    stale, mismatched, or absent when a shapefile is imported from a different
    source. If a raster transform is available we also convert georeferenced
    polygon coordinates back to pixel coordinates; otherwise we fall back to the
    original shape points directly.
    """
    if shape is None:
        return None

    pts = np.asarray(getattr(shape, "points", []), dtype=np.float64)
    if pts.size == 0:
        return None

    if raster_transform is not None and raster_bounds is not None:
        # If the polygon is georeferenced, convert geo -> pixel using the raster
        # transform, but only if the coordinates are spatially within the raster
        # bounds. This keeps imported results aligned with the viewer.
        try:
            px_pts = []
            for gx, gy in pts:
                col, row = (~raster_transform) * (gx, gy)
                if 0 <= col <= raster_bounds[2] and 0 <= row <= raster_bounds[3]:
                    px_pts.append((col, row))
            if px_pts:
                pts = np.asarray(px_pts, dtype=np.float64)
        except Exception:
            pass

    xs = pts[:, 0]
    ys = pts[:, 1]
    if xs.size == 0 or ys.size == 0:
        return None

    x1 = float(np.min(xs))
    y1 = float(np.min(ys))
    x2 = float(np.max(xs))
    y2 = float(np.max(ys))
    if x2 <= x1:
        x2 = x1 + 1.0
    if y2 <= y1:
        y2 = y1 + 1.0
    return [x1, y1, x2, y2]


def load_inference_result_from_shapefile(shp_path: Path, class_names: list = None, raster_path: Path = None):
    """Load a saved inference SHP output into an InferenceResult structure.

    The polygon geometry is treated as the source of truth because user-exported
    shapefiles may have stale or mismatched column values after editing or re-use.
    When a raster path is supplied, we also align georeferenced points to the
    raster's pixel grid so imported boxes land in the correct preview position.
    """
    shp_path = Path(shp_path)
    if not shp_path.is_file():
        raise FileNotFoundError(f"Shapefile hasil inference tidak ditemukan: {shp_path}")

    import shapefile
    raster_transform = None
    raster_bounds = None
    if raster_path is not None:
        raster_path = Path(raster_path)
        if raster_path.exists():
            try:
                with rasterio.open(raster_path) as src:
                    raster_transform = src.transform
                    raster_bounds = (0, 0, src.width, src.height)
            except Exception:
                raster_transform = None
                raster_bounds = None

    with shapefile.Reader(str(shp_path)) as shp:
        fields = [f[0] for f in shp.fields[1:]]
        boxes = []
        scores = []
        class_labels = []

        for record, shape in zip(shp.iterRecords(), shp.shapes()):
            values = dict(zip(fields, record))

            derived_box = _derive_box_from_shape(shape, raster_transform=raster_transform, raster_bounds=raster_bounds)
            if derived_box is not None:
                x1, y1, x2, y2 = derived_box
            else:
                x1 = float(values.get("x1_px", values.get("x1", 0.0)))
                y1 = float(values.get("y1_px", values.get("y1", 0.0)))
                x2 = float(values.get("x2_px", values.get("x2", 0.0)))
                y2 = float(values.get("y2_px", values.get("y2", 0.0)))
                if x2 <= x1:
                    x2 = x1 + 1.0
                if y2 <= y1:
                    y2 = y1 + 1.0

            boxes.append([x1, y1, x2, y2])
            scores.append(float(values.get("confidence", values.get("score", 0.0))))
            label = str(values.get("kelas", values.get("class", values.get("label", "sawit"))))
            class_labels.append(label or "sawit")

    if not boxes:
        return InferenceResult(
            boxes=np.zeros((0, 4), dtype=np.float32),
            scores=np.zeros((0,), dtype=np.float32),
            classes=np.zeros((0,), dtype=np.int32),
            class_names=[] if class_names is None else list(class_names),
            shp_path=shp_path,
        )

    unique_names = []
    for label in class_labels:
        if label not in unique_names:
            unique_names.append(label)
    if class_names is not None:
        unique_names = list(class_names)
    classes = np.array([unique_names.index(label) for label in class_labels], dtype=np.int32)

    return InferenceResult(
        boxes=np.asarray(boxes, dtype=np.float32),
        scores=np.asarray(scores, dtype=np.float32),
        classes=classes,
        class_names=unique_names,
        shp_path=shp_path,
    )


# ============================================================
# DATABASE (SQLITE) -- OPSIONAL, HANYA DIPAKAI KALAU db_path DIISI
# ============================================================
# Dibuat ringan & terpisah dari alur utama run() supaya TIDAK mengubah
# perilaku inferensi sama sekali kalau GUI belum butuh riwayat/registrasi
# model. Dua hal yang dicatat di sini: (1) riwayat tiap run (tabel 'runs',
# KF-16), dan (2) model itu sendiri disimpan UTUH beserta deskripsi
# tujuan/kegunaannya (tabel 'models', KF-15/KNF-11) -- bukan cuma path,
# supaya model tetap bisa dipilih & dimuat dari database walau file
# aslinya sudah dipindah/dihapus. Audit trail hasil koreksi TIDAK
# disimpan di sini -- itu ada di shapefile hasil koreksi sendiri
# (KNF-09), supaya tidak ada dua sumber kebenaran yang bisa tidak sinkron.

def _ensure_db(db_path: Path):
    import sqlite3
    con = sqlite3.connect(str(db_path))
    con.execute("""
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            raster_path TEXT, model_name TEXT, n_tiles INTEGER,
            n_detections INTEGER, elapsed_seconds REAL,
            shp_path TEXT, created_at TEXT
        )
    """)
    # KF-15 / KNF-11: model disimpan UTUH di database (isi file .pt dan isi
    # band_stats.json sebagai konten, bukan sekadar path ke disk), supaya
    # model tidak bergantung pada lokasi file aslinya -- tetap bisa dipilih
    # & dimuat dari database walau file sumbernya sudah dipindah/dihapus.
    con.execute("""
        CREATE TABLE IF NOT EXISTS models (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            model_name TEXT UNIQUE,
            pt_file BLOB,
            band_stats_json TEXT,
            description TEXT,
            created_at TEXT
        )
    """)
    con.commit()
    return con


def log_run_to_sqlite(db_path: Path, raster_path: Path, model_name: str,
                       n_tiles: int, n_detections: int, elapsed_seconds: float,
                       shp_path: Path) -> int:
    """Simpan satu baris riwayat run inference. Return run_id."""
    from datetime import datetime
    con = _ensure_db(Path(db_path))
    cur = con.execute(
        "INSERT INTO runs (raster_path, model_name, n_tiles, n_detections, "
        "elapsed_seconds, shp_path, created_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (str(raster_path), model_name, n_tiles, n_detections, elapsed_seconds,
         str(shp_path), datetime.now().isoformat(timespec="seconds")),
    )
    con.commit()
    run_id = cur.lastrowid
    con.close()
    return run_id


def register_model_in_sqlite(db_path: Path, model_name: str, pt_path: Path,
                              band_stats_path: Path, description: str = "") -> bool:
    """KF-15 -- dipanggil GUI saat pengguna mengimpor model baru (.pt +
    band_stats.json terkait). Membaca ISI file .pt dan band_stats.json lalu
    menyimpannya UTUH ke SQLite (bukan hanya path-nya), beserta deskripsi
    tujuan/kegunaan yang diminta ke pengguna saat impor -- sehingga model
    bisa dipilih & dimuat dari daftar di database tanpa perlu menelusuri
    atau bergantung pada file aslinya di disk."""
    from datetime import datetime
    try:
        with open(pt_path, "rb") as f:
            pt_bytes = f.read()
        with open(band_stats_path, "r") as f:
            band_stats_json = json.dumps(json.load(f))

        con = _ensure_db(Path(db_path))
        con.execute("""
            INSERT OR REPLACE INTO models (model_name, pt_file, band_stats_json, description, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (model_name, pt_bytes, band_stats_json, description or "",
              datetime.now().isoformat(timespec="seconds")))
        con.commit()
        con.close()
        return True
    except Exception as e:
        print(f"Gagal mendaftarkan model ke database: {e}")
        return False


def load_model_from_sqlite(db_path: Path, model_name: str) -> dict:
    """KF-15 -- ambil model dari database untuk dipakai inferensi. Isi file
    .pt ditulis ke file temporer (YOLO/ultralytics butuh path file, bukan
    bytes di memori) dan band_stats.json diparse langsung jadi dict.
    Deskripsi tujuan ikut dikembalikan supaya GUI bisa menampilkannya
    kembali ke pengguna setiap kali model ini dimuat.

    Return dict: {pt_path (temp file), band_stats (dict), description (str)}.
    """
    import tempfile
    import os

    con = _ensure_db(Path(db_path))
    cur = con.cursor()
    cur.execute(
        "SELECT pt_file, band_stats_json, description FROM models WHERE model_name = ?",
        (model_name,),
    )
    row = cur.fetchone()
    con.close()
    if not row:
        raise ValueError(f"Model '{model_name}' tidak ditemukan di database.")

    pt_bytes, band_stats_json, description = row
    temp_dir = tempfile.gettempdir()
    temp_pt_path = os.path.join(temp_dir, f"kanopai_model_{model_name}.pt")
    with open(temp_pt_path, "wb") as f:
        f.write(pt_bytes)

    raw_stats = json.loads(band_stats_json)
    band_stats = {int(k): v for k, v in raw_stats.items()}

    return {"pt_path": Path(temp_pt_path), "band_stats": band_stats, "description": description or ""}


def list_models_in_sqlite(db_path: Path) -> list:
    """Mengembalikan daftar model yang terdaftar di database, masing-masing
    {model_name, description}, buat ditampilkan sebagai daftar pilihan di
    GUI (tanpa perlu menarik isi file .pt yang berukuran besar)."""
    try:
        con = _ensure_db(Path(db_path))
        cur = con.cursor()
        cur.execute("SELECT model_name, description FROM models ORDER BY model_name")
        models = [{"model_name": row[0], "description": row[1] or ""} for row in cur.fetchall()]
        con.close()
        return models
    except Exception:
        return []


# ============================================================
# ENGINE
# ============================================================
class InferenceEngine:
    """
    Bungkus semua logic v2 script jadi satu class.
    log_fn(str) dan progress_fn(current, total) dipanggil dari worker thread GUI.
    should_cancel() dipanggil berkala; kalau True, proses dihentikan lewat CancelledError.
    """

    STRETCH_LOWER_PCT = 1.0
    STRETCH_UPPER_PCT = 99.0

    def __init__(self, model_path: str, band_stats_path: str,
                 log_fn=print, progress_fn=None, should_cancel=None):
        self.model_path = Path(model_path)
        self.band_stats_path = Path(band_stats_path)
        self.log_fn = log_fn
        self.progress_fn = progress_fn or (lambda cur, total: None)
        self.should_cancel = should_cancel or (lambda: False)

        self.model = None
        self.band_stats = None
        self.device = None

    def load(self):
        if not self.model_path.is_file():
            raise FileNotFoundError(f"Model tidak ditemukan: {self.model_path}")
        if not self.band_stats_path.is_file():
            raise FileNotFoundError(f"band_stats tidak ditemukan: {self.band_stats_path}")

        # Deteksi & pakai GPU secara EKSPLISIT -- kalau tidak diset manual,
        # ultralytics kadang jatuh ke CPU tanpa terlihat jelas di log.
        if torch.cuda.is_available():
            self.device = 0
            gpu_name = torch.cuda.get_device_name(0)
            self.log_fn(f"GPU terdeteksi: {gpu_name}. Inference akan pakai GPU.")
        else:
            self.device = "cpu"
            self.log_fn("[PERINGATAN] GPU/CUDA TIDAK terdeteksi oleh torch. "
                         "Inference akan jalan di CPU dan JAUH lebih lambat. "
                         "Cek instalasi torch+CUDA kamu.")

        self.log_fn(f"Memuat model: {self.model_path.name} ...")
        self.model = YOLO(str(self.model_path))

        # WORKAROUND bug ultralytics: saat predict() dipanggil pertama kali,
        # BasePredictor.stream_inference() otomatis melakukan "warmup" dengan
        # tensor dummy yang channel-nya DI-HARDCODE 3 (lihat
        # ultralytics/engine/predictor.py -- tidak menyesuaikan channel count
        # model asli). Untuk model 7-channel kita, warmup itu crash duluan
        # ("expected input to have 7 channels, but got 3") sebelum tile asli
        # sempat diproses. Solusinya: bikin predictor secara manual di sini,
        # lalu tandai done_warmup=True supaya baris warmup yang salah channel
        # itu DILEWATI oleh ultralytics -- forward pass sesungguhnya nanti
        # tetap jalan normal karena tile yang kita kirim sendiri (lewat
        # _flush_batch -> model.predict(source=tile_batch, ...)) sudah punya
        # channel count yang benar (7). Efek sampingnya cuma inference
        # tile PERTAMA sedikit lebih lambat (CUDA belum di-"panaskan"),
        # tidak ada perubahan hasil deteksi sama sekali.
        try:
            predictor_cls = self.model._smart_load("predictor")
            predictor_overrides = {
                **self.model.overrides,
                "conf": 0.25, "batch": 1, "save": False, "mode": "predict",
            }
            self.model.predictor = predictor_cls(
                overrides=predictor_overrides, _callbacks=self.model.callbacks
            )
            self.model.predictor.setup_model(model=self.model.model, verbose=False)
            self.model.predictor.done_warmup = True
        except Exception as e:
            self.log_fn(
                f"[PERINGATAN] Gagal bypass warmup ultralytics ({e}); "
                f"kalau model ini bukan 7-channel, ini biasanya aman "
                f"diabaikan."
            )

        names = getattr(self.model, "names", None)
        if names is None:
            self.class_names = None
        elif isinstance(names, dict):
            self.class_names = [str(v) for v in names.values()]
        elif isinstance(names, (list, tuple, np.ndarray)):
            self.class_names = [str(v) for v in names]
        else:
            self.class_names = None
        self.band_stats = load_band_stats(self.band_stats_path)
        self.log_fn(f"Model & band stats siap ({len(self.band_stats)} slot).")

    def run(self, raster_path: str, conf: float = 0.25, tile_size: int = 640,
            overlap: int = 64, iou_threshold: float = 0.5,
            output_dir: str = None, batch_size: int = 4, out_name: str = None,
            aoi_shp_path: str = None, exclude_shp_path: str = None,
            aoi_polygons_px: list = None, exclude_polygons_px: list = None,
            db_path: str = None, manual_band_mapping: dict = None,
            enable_adaptive_fallback: bool = False) -> InferenceResult:
        """
        manual_band_mapping: FITUR OPSIONAL. Dict {slot_training: band_input}
        untuk override auto-detect band mapping -- dipakai kalau pengguna
        mau menentukan sendiri pasangan band (mis. hasil auto-detect
        meleset karena mean band kebetulan mirip). Kalau tidak diisi
        (None, default), perilaku lama tetap jalan: auto-detect sesuai
        skema band_stats (1-to-1 / multiref / adaptif).
        
        batch_size default diturunkan ke 4 (dari 8) -- target device Windows
        10 dengan RAM 8GB, batch besar berisiko OOM terutama saat model
        gabungan (7 channel) dipakai. GUI tetap boleh menaikkan ini kalau
        device penggunanya lebih besar.

        aoi_shp_path / exclude_shp_path: FITUR OPSIONAL. Path shapefile
        polygon AOI (area yang mau diproses) dan/atau exclude area (area
        yang mau dikecualikan, mis. jalan/badan air/bangunan di tengah
        kebun). Kalau tidak diisi, seluruh raster diproses seperti biasa
        (tidak mengubah perilaku lama).

        db_path: FITUR OPSIONAL. Kalau diisi, riwayat run ini dicatat ke
        SQLite (lihat log_run_to_sqlite). Tidak wajib dipakai.

        enable_adaptive_fallback: FITUR OPSIONAL, default False (OFF).
        Dipakai HANYA kalau raster datang dari sensor yang jumlah/karakteristik
        band-nya tidak cocok dengan referensi manapun di band_stats DAN
        manual_band_mapping tidak diisi -- lihat auto_detect_band_mapping_adaptive
        di bawah. Kalau False (default) dan sensor tak dikenali, run() akan
        berhenti dengan ValueError yang jelas (bukan menebak diam-diam),
        supaya pengguna sadar dan bisa pilih matching manual atau menyalakan
        mode ini secara sengaja. Mode ini adalah jaring pengaman sementara,
        BUKAN pengganti fine-tuning model untuk sensor baru tersebut.
        """
        """
        Deduplikasi duplikat tile-boundary pakai pendekatan "kepemilikan
        wilayah tile" (lihat filter_by_tile_ownership), BUKAN tebak-tebakan
        jarak/IoU antar box. Setiap titik di raster cuma "dimiliki" oleh satu
        tile (ditentukan dari posisi tile itu sendiri), sehingga duplikat di
        zona overlap otomatis tersingkir tanpa perlu menaksir "ini pohon sama
        atau pohon beda" dari geometrinya -- ini yang dulu gagal di kanopi
        kecil & rapat (2 pohon beda salah dianggap 1 pohon yang sama karena
        kebetulan berdekatan).
        """
        import time
        _t_start = time.perf_counter()

        if self.model is None:
            self.load()

        raster_path = Path(raster_path)
        if not raster_path.is_file():
            raise FileNotFoundError(f"Raster tidak ditemukan: {raster_path}")

        self.log_fn(f"Membuka raster: {raster_path.name}")
        # GDAL_CACHEMAX dibatasi (256MB) supaya aman di device RAM 8GB dan
        # saat raster berformat BigTIFF (>4GB) -- rasterio/GDAL sudah baca
        # BigTIFF secara native lewat windowed read yang sudah dipakai di
        # pipeline ini, jadi tidak perlu penanganan khusus selain batasan
        # cache ini agar GDAL tidak menahan buffer besar di RAM.
        with rasterio.Env(GDAL_CACHEMAX=256, GDAL_TIFF_INTERNAL_MASK=True), \
             rasterio.open(raster_path) as src:
            width, height, n_bands = src.width, src.height, src.count
            self.log_fn(f"Ukuran raster: {width} x {height} px, {n_bands} band")

            raster_dtype = str(src.dtypes[0])
            is_uint8_input = raster_dtype == "uint8"

            expected_n_bands = len(self.band_stats)
            multiref = is_multiref_schema(self.band_stats)

            if manual_band_mapping is not None:
                self.log_fn("Mode manual: memakai band mapping dari pengguna (auto-detect dilewati).")
                band_mapping = validate_manual_band_mapping(
                    manual_band_mapping, self.band_stats, n_bands, log=self.log_fn
                )
            elif n_bands == expected_n_bands and not multiref:
                band_mapping = {b: b for b in range(1, expected_n_bands + 1)}
                self.log_fn(f"Band lengkap ({n_bands}). Mapping 1-to-1.")
            elif multiref:
                self.log_fn(f"Model gabungan terdeteksi. Mencocokkan {n_bands} band input -> {expected_n_bands} slot...")
                band_mapping = auto_detect_band_mapping_multiref(src, self.band_stats, log=self.log_fn)
            elif not enable_adaptive_fallback:
                raise ValueError(
                    f"Sensor input tidak dikenali sistem: jumlah band raster "
                    f"({n_bands}) tidak sama dengan slot band training "
                    f"({expected_n_bands}), dan tidak cocok skema model gabungan. "
                    f"Pilih salah satu: (1) isi manual_band_mapping untuk "
                    f"memetakan band secara manual, atau (2) set "
                    f"enable_adaptive_fallback=True untuk memakai jaring "
                    f"pengaman sementara (parameter normalisasi dihitung "
                    f"otomatis dari raster ini). Opsi (2) bukan pengganti "
                    f"fine-tuning model untuk sensor ini."
                )
            else:
                self.log_fn(
                    f"[PERINGATAN] Jumlah band beda ({n_bands} vs {expected_n_bands}) "
                    f"dan sensor tidak dikenali. Memakai fallback adaptif "
                    f"(parameter normalisasi dihitung dari raster ini, bukan "
                    f"dipinjam dari sensor lain) -- ini jaring pengaman "
                    f"sementara, bukan pengganti fine-tuning."
                )
                band_mapping = auto_detect_band_mapping_adaptive(
                    src, self.band_stats, self.STRETCH_LOWER_PCT,
                    self.STRETCH_UPPER_PCT, log=self.log_fn
                )

            # --- AOI / exclude area (opsional) ---
            aoi_polys, exclude_polys = [], []
            if aoi_shp_path:
                aoi_polys.extend(load_polygons_px(Path(aoi_shp_path), src.transform))
            if aoi_polygons_px:
                aoi_polys.extend(polygons_from_pixel_coords(aoi_polygons_px))
            if aoi_polys:
                self.log_fn(f"AOI dimuat: {len(aoi_polys)} polygon.")
            else:
                aoi_polys = None

            if exclude_shp_path:
                exclude_polys.extend(load_polygons_px(Path(exclude_shp_path), src.transform))
            if exclude_polygons_px:
                exclude_polys.extend(polygons_from_pixel_coords(exclude_polygons_px))
            if exclude_polys:
                self.log_fn(f"Exclude area dimuat: {len(exclude_polys)} polygon.")
            else:
                exclude_polys = None

            # Crop ke bounding-box AOI, lalu generate tile di region crop
            if aoi_polys:
                crop_x, crop_y, crop_w, crop_h = compute_aoi_crop_region(
                    aoi_polys, width, height
                )
                self.log_fn(
                    f"AOI crop region: x={crop_x}, y={crop_y}, "
                    f"w={crop_w}, h={crop_h} (crop → pad → scan)"
                )
                local_windows = generate_tile_windows(crop_w, crop_h, tile_size, overlap)
                windows = [
                    (x + crop_x, y + crop_y, w, h)
                    for x, y, w, h in local_windows
                ]
            else:
                windows = generate_tile_windows(width, height, tile_size, overlap)

            if aoi_polys:
                before = len(windows)
                windows = [w for w in windows if tile_intersects_aoi(*w, aoi_polys)]
                self.log_fn(f"Filter AOI: {before} tile -> {len(windows)} tile relevan diproses.")
            windows_by_id = {idx: w for idx, w in enumerate(windows, start=1)}
            total = len(windows)
            self.log_fn(f"Akan diproses {total} tile ({tile_size}x{tile_size}, overlap {overlap}px)")

            all_boxes, all_scores, all_classes, all_tile_ids = [], [], [], []

            def _prepare_tile(x_off, y_off, w, h):
                """Baca & stretch satu tile dari raster (I/O + CPU, TIDAK menyentuh GPU)."""
                tile_chw = np.zeros((expected_n_bands, h, w), dtype=np.uint8)
                window = rasterio.windows.Window(x_off, y_off, w, h)

                for target_b in range(1, expected_n_bands + 1):
                    entry = band_mapping.get(target_b)
                    if entry is None:
                        continue

                    if isinstance(entry, dict):
                        input_b_idx = entry["input_band"]
                        fallback_p_low = entry.get("p_low")
                        fallback_p_high = entry.get("p_high")
                    else:
                        input_b_idx = entry
                        stats = self.band_stats.get(target_b, {})
                        fallback_p_low = stats.get("p_low")
                        fallback_p_high = stats.get("p_high")

                    data = src.read(input_b_idx, window=window)

                    if is_uint8_input:
                        stretched = data.astype(np.uint8)
                    else:
                        valid_pixels = data[data > 0]
                        if len(valid_pixels) > (w * h * 0.05):
                            p_low, p_high = np.percentile(valid_pixels, (self.STRETCH_LOWER_PCT, self.STRETCH_UPPER_PCT))
                        elif fallback_p_low is not None and fallback_p_high is not None:
                            p_low, p_high = fallback_p_low, fallback_p_high
                        else:
                            p_low, p_high = 0, 255
                        stretched = stretch_band(data, p_low, p_high)

                    tile_chw[target_b - 1] = stretched

                tile_chw = apply_polygon_masks_to_tile(
                    tile_chw, x_off, y_off, w, h, aoi_polys, exclude_polys
                )
                tile_hwc = tile_chw.transpose(1, 2, 0)
                return pad_tile_for_inference(tile_hwc, target_size=tile_size)

            def _flush_batch(tile_batch, offset_batch):
                """Kirim satu batch tile sekaligus ke model -- GPU jauh lebih efisien
                diberi banyak gambar sekaligus daripada satu-satu."""
                results = self.model.predict(source=tile_batch, device=self.device,
                                              conf=conf, save=False, verbose=False)
                n_det_total = 0
                for r, (x_off, y_off, tile_idx) in zip(results, offset_batch):
                    if r.boxes is not None and len(r.boxes) > 0:
                        boxes_xyxy = r.boxes.xyxy.cpu().numpy()
                        scores = r.boxes.conf.cpu().numpy()
                        classes = r.boxes.cls.cpu().numpy()
                        boxes_xyxy[:, [0, 2]] += x_off
                        boxes_xyxy[:, [1, 3]] += y_off
                        all_boxes.append(boxes_xyxy)
                        all_scores.append(scores)
                        all_classes.append(classes)
                        all_tile_ids.append(np.full(len(scores), tile_idx, dtype=np.int32))
                        n_det_total += len(scores)
                return n_det_total

            tile_batch, offset_batch = [], []

            for idx, (x_off, y_off, w, h) in enumerate(windows, start=1):
                if self.should_cancel():
                    raise CancelledError("Dibatalkan oleh pengguna.")

                tile_batch.append(_prepare_tile(x_off, y_off, w, h))
                offset_batch.append((x_off, y_off, idx))

                is_last = (idx == total)
                if len(tile_batch) >= batch_size or is_last:
                    n_det = _flush_batch(tile_batch, offset_batch)
                    self.log_fn(f"[{idx}/{total}] batch selesai ({len(tile_batch)} tile), "
                                f"{n_det} objek di batch ini")
                    tile_batch, offset_batch = [], []
                    gc.collect()  # lepas memori tile-batch -- penting di RAM 8GB

                self.progress_fn(idx, total)

        result = InferenceResult()
        result.n_tiles = total
        if not all_boxes:
            self.log_fn("Tidak ada objek terdeteksi di seluruh raster.")
            result.elapsed_seconds = time.perf_counter() - _t_start
            return result

        all_boxes = np.concatenate(all_boxes, axis=0)
        all_scores = np.concatenate(all_scores, axis=0)
        all_classes = np.concatenate(all_classes, axis=0)
        all_tile_ids = np.concatenate(all_tile_ids, axis=0)

        self.log_fn(f"Total deteksi sebelum dedup: {len(all_boxes)}")

        # --- Tahap 1: Dedup duplikat tile-boundary via kepemilikan wilayah ---
        ownership_keep = filter_by_tile_ownership(
            all_boxes, all_tile_ids, windows_by_id, width, height, overlap
        )
        n_removed_ownership = int((~ownership_keep).sum())
        all_boxes = all_boxes[ownership_keep]
        all_scores = all_scores[ownership_keep]
        all_classes = all_classes[ownership_keep]
        self.log_fn(
            f"[Dedup] Kepemilikan wilayah tile: {n_removed_ownership} deteksi "
            f"disingkirkan (titik tengahnya jatuh di zona overlap milik tile "
            f"tetangga, bukan tile ini). Sisa: {len(all_boxes)}."
        )

        # --- Tahap 2: NMS IoU standar (bersihkan sisa duplikat DALAM 1 tile) ---
        keep_idx = nms_global(all_boxes, all_scores, classes=all_classes, iou_threshold=iou_threshold)
        final_boxes = all_boxes[keep_idx]
        final_scores = all_scores[keep_idx]
        final_classes = all_classes[keep_idx]
        self.log_fn(f"Total deteksi setelah NMS: {len(final_boxes)}")

        # --- Tahap 3: Filter AOI/exclude area (opsional) ---
        if aoi_polys or exclude_polys:
            aoi_keep = filter_boxes_by_aoi(final_boxes, aoi_polys, exclude_polys)
            n_removed_aoi = int((~aoi_keep).sum())
            final_boxes = final_boxes[aoi_keep]
            final_scores = final_scores[aoi_keep]
            final_classes = final_classes[aoi_keep]
            self.log_fn(
                f"[AOI/Exclude] {n_removed_aoi} deteksi disingkirkan (di luar AOI "
                f"atau di dalam exclude area). Sisa: {len(final_boxes)}."
            )

        model_stem = self.model_path.stem
        # Hasil file hanya dibuat melalui tombol Export Shapefiles.
        # Inference tetap menyiapkan preview di memori untuk tampilan UI.
        class_names = getattr(self, "class_names", None)

        self.log_fn("Membuat preview visual di memori...")
        preview_bgr, preview_scale = build_preview_bgr(raster_path, final_boxes, final_scores,
                                         self.STRETCH_LOWER_PCT, self.STRETCH_UPPER_PCT,
                                         classes=final_classes)

        result.boxes = final_boxes
        result.scores = final_scores
        result.classes = final_classes
        result.class_names = self.class_names if getattr(self, "class_names", None) is not None else class_names
        result.class_counts = summarize_class_counts(final_classes, result.class_names)
        result.preview_bgr = preview_bgr
        result.preview_scale = preview_scale
        result.aoi_polygons_px = aoi_polys
        result.exclude_polygons_px = exclude_polys
        result.elapsed_seconds = time.perf_counter() - _t_start

        # --- Logging SQLite (opsional) ---
        if db_path:
            try:
                run_id = log_run_to_sqlite(
                    db_path, raster_path, model_stem, total, len(final_boxes),
                    result.elapsed_seconds, None,
                )
                result.run_id = run_id
                self.log_fn(f"Riwayat run dicatat ke database (run_id={run_id}).")
            except Exception as e:
                # Jangan sampai kegagalan logging DB menggagalkan seluruh inference
                self.log_fn(f"[PERINGATAN] Gagal mencatat ke database: {e}")

        return result
